#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Injeta pedidos históricos 2024/2025/2026 do arquivo /3MV/BI/Pedidos_Historico/Pedidos_2024_2025_2026.xlsx
no dados.json. Permite comparativo granular por cliente×indústria×ano.

Estrutura adicionada em dados.json:
  d['historico_pedidos'] = {
    'planilha_path': '...',
    'planilha_mtime': 'YYYY-MM-DD HH:MM',
    'total_linhas': N,
    'por_cliente_industria_ano': {  # cliente → industria → ano → {valor, faturado, saldo, qtd_pedidos}
        'CLIENTE X': {'IND Y': {'2024': {...}, '2025': {...}, '2026': {...}}}
    },
    'mensal_por_industria': {  # industria → ano → mes → valor (mês 1-12)
        'EMBELLEZE': {'2025': [v1..v12], '2026': [v1..v12], '2024': [v1..v12]}
    },
    'top_clientes_25_26': [...]  # lista pré-ordenada pra display rápido
  }
"""
import json
import sys
import os
from pathlib import Path
from datetime import datetime
try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install',
                           '--break-system-packages', '--quiet', 'openpyxl'])
    from openpyxl import load_workbook


def _gd_base():
    candidates = [
        Path.home() / 'Library/CloudStorage/GoogleDrive-marcelo@3mvrepresentacao.com/Meu Drive',
        Path(__file__).resolve().parent.parent.parent.parent.parent,
        Path(os.environ.get('GD_MEU_DRIVE', '/dev/null')),
    ]
    for c in candidates:
        try:
            if (c / '3MV/BI/Pedidos_Historico/Pedidos_2024_2025_2026.xlsx').exists():
                return c
        except Exception:
            continue
    return candidates[0]


GD = _gd_base()
XLSX = GD / '3MV/BI/Pedidos_Historico/Pedidos_2024_2025_2026.xlsx'
DADOS_JSON = Path(__file__).parent / 'dados.json'


def _val(v):
    if v is None: return ''
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)): return v
    return str(v).strip()


def _to_num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try:
        return float(str(v).replace(',', '.').replace('R$', '').strip())
    except Exception:
        return 0.0


def _ano_mes(v):
    """Retorna (ano, mes) a partir de Dt. Venda. Se não houver, None."""
    if isinstance(v, datetime):
        return v.year, v.month
    if isinstance(v, str) and len(v) >= 10:
        try:
            d = datetime.strptime(v[:10], '%Y-%m-%d')
            return d.year, d.month
        except Exception:
            pass
    return None, None


def main():
    print(f'\n→ Lendo histórico de pedidos: {XLSX}')
    if not XLSX.exists():
        print(f'  ⚠ {XLSX} não existe — pulando injeção')
        return 0

    wb = load_workbook(XLSX, data_only=True)
    sheet = wb.active   # 1ª aba ('Pedidos')
    print(f'  Aba: {sheet.title} ({sheet.max_row} linhas)')

    por_ci_ano = {}     # cliente → industria → ano → {valor, faturado, saldo, qtd}
    mensal_ind = {}     # industria → ano → [12 meses de valor]
    todos_pedidos = []  # raw data pra outras consultas

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:  # sem cliente, pula
            continue
        try:
            tipo = _val(row[0])
            cliente = _val(row[1]).upper().strip() if row[1] else ''
            industria = _val(row[2]).upper().strip() if len(row) > 2 and row[2] else ''
            numero = row[3] if len(row) > 3 else None
            valor = _to_num(row[6]) if len(row) > 6 else 0
            valor_fat = _to_num(row[7]) if len(row) > 7 else 0
            saldo = _to_num(row[8]) if len(row) > 8 else 0
            dt_venda = row[9] if len(row) > 9 else None
            status = _val(row[10]) if len(row) > 10 else ''
            ano, mes = _ano_mes(dt_venda)
            if not ano or not industria or not cliente:
                continue

            # Pedido (não duplicar quando há múltiplas linhas — usar tipo='Pedido')
            if tipo and 'PEDIDO' not in tipo.upper():
                continue

            ano_str = str(ano)
            por_ci_ano.setdefault(cliente, {}).setdefault(industria, {}).setdefault(ano_str, {
                'valor': 0, 'faturado': 0, 'saldo': 0, 'qtd_pedidos': 0
            })
            d = por_ci_ano[cliente][industria][ano_str]
            d['valor'] += valor
            d['faturado'] += valor_fat
            d['saldo'] += saldo
            d['qtd_pedidos'] += 1

            mensal_ind.setdefault(industria, {}).setdefault(ano_str, [0]*12)
            if 1 <= mes <= 12:
                mensal_ind[industria][ano_str][mes-1] += valor

            todos_pedidos.append({
                'cliente': cliente, 'industria': industria, 'numero': numero,
                'valor': valor, 'faturado': valor_fat, 'saldo': saldo,
                'ano': ano, 'mes': mes, 'status': status,
            })
        except Exception as e:
            continue

    # Top clientes 25 vs 26 — pré-ordenado
    top_clientes = []
    for cli, inds in por_ci_ano.items():
        v25 = sum(d.get('2025', {}).get('valor', 0) for d in inds.values())
        v26 = sum(d.get('2026', {}).get('valor', 0) for d in inds.values())
        v24 = sum(d.get('2024', {}).get('valor', 0) for d in inds.values())
        if v24 + v25 + v26 > 0:
            dpct = ((v26/v25 - 1)*100) if v25 > 0 else None
            top_clientes.append({
                'cliente': cli, 'fat_24': v24, 'fat_25': v25, 'fat_26': v26,
                'delta_pct_26x25': dpct,
                'industrias': sorted(inds.keys()),
            })
    top_clientes.sort(key=lambda x: x['fat_26'], reverse=True)

    # Carrega dados.json e injeta
    if not DADOS_JSON.exists():
        print(f'  ⚠ dados.json não existe — pulando')
        return 0

    d = json.loads(DADOS_JSON.read_text(encoding='utf-8'))
    try:
        mtime = datetime.fromtimestamp(XLSX.stat().st_mtime).strftime('%Y-%m-%d %H:%M')
    except Exception:
        mtime = ''

    d['historico_pedidos'] = {
        '_doc': 'Lido de /3MV/BI/Pedidos_Historico/Pedidos_2024_2025_2026.xlsx',
        '_atualizado_em': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'planilha_path': str(XLSX),
        'planilha_mtime': mtime,
        'total_linhas': len(todos_pedidos),
        'por_cliente_industria_ano': por_ci_ano,
        'mensal_por_industria': mensal_ind,
        'top_clientes': top_clientes[:300],   # limita pra payload
    }

    DADOS_JSON.write_text(json.dumps(d, ensure_ascii=False), encoding='utf-8')
    print(f'  ✓ Injetado {len(todos_pedidos)} pedidos · {len(por_ci_ano)} clientes · {len(mensal_ind)} indústrias')
    print(f'  Top 5 clientes 26: {[t["cliente"] for t in top_clientes[:5]]}')
    return 0


if __name__ == '__main__':
    sys.exit(main() or 0)
