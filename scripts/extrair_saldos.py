#!/usr/bin/env python3
"""
EXTRAIR SALDOS DE PEDIDOS — 3MV Representação

Conecta na API Suas Vendas (api.suasvendas.com/v2), baixa todos os pedidos
do ano-alvo (default 2026), identifica os pedidos com FATURAMENTO PARCIAL
e gera duas saídas:

  • JSON estruturado com os itens em saldo, agrupados por
    cliente / mês (data fatura) / produto.
  • Excel formatado para envio aos destinatários.

USO:
  python3 extrair_saldos.py \
      --config /caminho/config.json \
      --output-json /tmp/saldos.json \
      --output-xlsx /tmp/saldos.xlsx \
      [--ano 2026]

REGRA DE "FATURADO PARCIAL":
  Pedido (não excluído) que tem AO MENOS UM item com peit_qtde_faturada > 0
  E AO MENOS UM item com peit_qtde > peit_qtde_faturada.
  Essa definição é robusta contra mudanças no campo pedi_status.

CRÉDITOS / REUSO:
  Paginação híbrida (paginado + busca binária + GETs individuais) reaproveitada
  do projeto /Automações/comissoes/scripts/atualizar.py — necessária por causa
  do bug conhecido em /Pedido que para a paginação no id ~9009.
"""
from __future__ import annotations

import argparse
import calendar
import json
import sys
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

# ============================================================
# CONFIG / CREDENCIAIS
# ============================================================

DEFAULT_HOST = "https://api.suasvendas.com/v2"


def find_drive_root() -> Path:
    """Localiza a raiz do Google Drive do Marcelo (host real ou sandbox)."""
    candidatos = [
        Path.home() / "Library/CloudStorage/GoogleDrive-marcelo@3mvrepresentacao.com/Meu Drive",
        Path("/sessions/confident-festive-mendel/mnt/Meu Drive"),
        Path("/sessions/confident-festive-mendel/mnt"),
    ]
    import os
    if os.environ.get("GD_BASE"):
        candidatos.insert(0, Path(os.environ["GD_BASE"]))
    for c in candidatos:
        if (c / "3MV/Automações/comissoes/credenciais.json").exists():
            return c
        if (c / "3MV").exists():
            return c
    return candidatos[0]


def carregar_credenciais(path: Path) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"Credenciais não encontradas em: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def montar_headers(creds: dict) -> dict:
    token_raw = creds["authorization"].replace("Yoursoft ", "").strip()
    cliente = creds.get("cliente") or creds.get("Cliente") or "representacao3mv"
    return {
        "Token": token_raw,
        "Cliente": cliente,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }


# ============================================================
# CHAMADAS HTTP — paginação híbrida
# ============================================================

def _get(host: str, path: str, headers: dict, qs: str = "", timeout: int = 120) -> tuple[int, str]:
    sep = "?" if qs else ""
    url = f"{host}{path}{sep}{qs}"
    req = urllib.request.Request(url, headers=headers, method="GET")
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        body = resp.read(50_000_000).decode("utf-8", errors="replace")
        return resp.status, body


def _parse_json_lenient(body: str) -> Any:
    try:
        return json.loads(body)
    except json.JSONDecodeError:
        try:
            data, _ = json.JSONDecoder().raw_decode(body)
            return data
        except Exception:
            return None


def fetch_pagina(host: str, path: str, headers: dict, page: int, extra_qs: str = "") -> tuple[int, list]:
    qs_parts = [f"page={page}"]
    if extra_qs:
        qs_parts.append(extra_qs)
    qs = "&".join(qs_parts)
    status, body = _get(host, path, headers, qs=qs)
    if status != 200:
        return status, []
    data = _parse_json_lenient(body)
    if not isinstance(data, list):
        return status, []
    return status, data


def fetch_all(host: str, path: str, headers: dict, extra_qs: str = "", max_pages: int = 500, label: str = "") -> list:
    out: list = []
    page = 1
    label = label or path
    while page <= max_pages:
        try:
            status, data = fetch_pagina(host, path, headers, page, extra_qs)
        except Exception as e:
            print(f"  ! erro pág {page} de {label}: {e}")
            break
        if status != 200 or not data:
            if page == 1 and status != 200:
                print(f"  status {status} na primeira página de {label}, abortando")
            break
        out.extend(data)
        qtde = data[0].get("qtde_paginas", 1) if isinstance(data[0], dict) else 1
        if page == 1 or page % 10 == 0 or page == qtde:
            print(f"  {label} pág {page}/{qtde} (acumulado {len(out)})")
        if page >= qtde:
            break
        page += 1
    return out


def fetch_pedido_individual(host: str, headers: dict, pid: int) -> dict | None:
    return fetch_individual(host, headers, "/Pedido", pid)


def fetch_individual(host: str, headers: dict, base_path: str, rid: int) -> dict | None:
    """GET genérico em /<recurso>/{id}. Retorna o dict ou None em caso de 4xx/erro."""
    try:
        status, body = _get(host, f"{base_path}/{rid}", headers, timeout=30)
    except urllib.error.HTTPError:
        return None
    except Exception:
        return None
    if status != 200:
        return None
    data = _parse_json_lenient(body)
    if isinstance(data, list):
        return data[0] if data else None
    return data if isinstance(data, dict) else None


def descobrir_id_max(host: str, headers: dict, id_min: int, max_busca: int = 10000) -> int:
    """Busca exponencial + binária pelo maior pedi_id que ainda existe."""
    lo = id_min
    passo = 100
    while passo <= max_busca:
        cand = lo + passo
        if fetch_pedido_individual(host, headers, cand) is not None:
            lo = cand
            passo *= 2
        else:
            break
    hi = min(lo + passo, lo + max_busca)
    while lo + 1 < hi:
        mid = (lo + hi) // 2
        if fetch_pedido_individual(host, headers, mid) is not None:
            lo = mid
        else:
            hi = mid
    return lo


def fetch_pedidos_completo(host: str, headers: dict) -> list:
    """
    Workaround do bug de paginação: paginado + busca binária + GETs individuais
    para os pedidos com id maior que o limite do paginado.
    """
    # FIX 2026-05-04 v3: usar /Pedido sem filtros — o ?excluido=0 retorna
    # pedi_itens com peit_qtde_faturada zerado (bug da API). Filtrar excluido em Python.
    print("  fase 1: paginado /Pedido (sem filtros — peit_qtde_faturada correto) ...")
    pedidos = fetch_all(host, "/Pedido", headers, label="/Pedido")
    if not pedidos:
        print("  fallback: paginado /Pedido?excluido=0 ...")
        pedidos = fetch_all(host, "/Pedido", headers, label="/Pedido")
    ids_set = {p.get("pedi_id") for p in pedidos if p.get("pedi_id") is not None}
    id_max_pag = max(ids_set) if ids_set else 0
    print(f"  paginado: {len(pedidos)} pedidos, pedi_id_max={id_max_pag}")

    print(f"  fase 2: descobrindo último pedi_id real (busca binária a partir de {id_max_pag}) ...")
    id_max_real = descobrir_id_max(host, headers, id_max_pag)
    print(f"  id_max_real = {id_max_real}")

    if id_max_real > id_max_pag:
        faltantes = list(range(id_max_pag + 1, id_max_real + 1))
        print(f"  fase 3: baixando {len(faltantes)} pedidos individuais ({id_max_pag + 1} → {id_max_real}) ...")
        # FASE 9.63: paralelismo (10 threads) — corta tempo de ~10min p/ ~1min
        adicionados = 0
        gaps = 0
        a_buscar = [pid for pid in faltantes if pid not in ids_set]
        with ThreadPoolExecutor(max_workers=10) as ex:
            futs = {ex.submit(fetch_pedido_individual, host, headers, pid): pid for pid in a_buscar}
            for fut in as_completed(futs):
                p = fut.result()
                if p is None:
                    gaps += 1
                    continue
                pedidos.append(p)
                ids_set.add(futs[fut])
                adicionados += 1
                if adicionados % 100 == 0:
                    print(f"    +{adicionados}/{len(a_buscar)} (gaps {gaps})")
        print(f"  fase 3 ok: +{adicionados} pedidos individuais (gaps {gaps})")

    # FIX 2026-05-04 v4: a paginação do /Pedido tem bug — alguns IDs DENTRO do
    # range [1..id_max_pag] não aparecem em nenhuma página. Aqui descobrimos
    # quais estão faltando e buscamos individualmente.
    # Estratégia: limitar a busca aos últimos 2000 IDs (ano corrente) pra
    # evitar fetch de pedidos antigos (anos passados).
    busca_min = max(1, id_max_real - 2000)
    print(f"  fase 4: completando IDs faltantes em [{busca_min}, {id_max_real}] ...")
    faltantes_dentro = [i for i in range(busca_min, id_max_real + 1) if i not in ids_set]
    print(f"  fase 4: {len(faltantes_dentro)} IDs ausentes da paginação dentro do range")
    # FASE 9.63: paralelismo (10 threads)
    add4 = 0
    gaps4 = 0
    with ThreadPoolExecutor(max_workers=10) as ex:
        futs = {ex.submit(fetch_pedido_individual, host, headers, pid): pid for pid in faltantes_dentro}
        for fut in as_completed(futs):
            p = fut.result()
            if p is None:
                gaps4 += 1
                continue
            pedidos.append(p)
            ids_set.add(futs[fut])
            add4 += 1
            if add4 % 200 == 0:
                print(f"    fase 4 +{add4}/{len(faltantes_dentro)} (gaps {gaps4})")
    print(f"  fase 4 ok: +{add4} pedidos recuperados (gaps {gaps4})")
    return pedidos


# ============================================================
# AUXILIARES — datas / formatos
# ============================================================

def parse_iso(s) -> datetime | None:
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "").split(".")[0])
    except Exception:
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d")
        except Exception:
            return None


MESES_PT = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
            "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]


def to_float(v, default: float = 0.0) -> float:
    try:
        return float(v) if v not in (None, "") else default
    except (TypeError, ValueError):
        return default


def to_int(v, default=None):
    try:
        return int(v) if v not in (None, "") else default
    except (TypeError, ValueError):
        return default


# ============================================================
# CORE — extrai itens em saldo
# ============================================================

def construir_mapa_clientes(host: str, headers: dict) -> dict[int, str]:
    print("\n[1/3] Baixando clientes ...")
    clientes = fetch_all(host, "/Cliente", headers, label="/Cliente")
    mapa: dict[int, str] = {}
    for c in clientes:
        cid = to_int(c.get("cont_id"))
        if cid is None:
            continue
        nome = (c.get("cont_nome_fantasia") or c.get("cont_razao_social") or "").strip()
        if nome:
            mapa[cid] = nome
    print(f"  {len(mapa)} clientes mapeados")
    return mapa


def construir_mapa_industrias(host: str, headers: dict) -> dict[int, str]:
    print("\n[2/3] Baixando indústrias ...")
    industrias = fetch_all(host, "/Industria", headers, label="/Industria")
    mapa: dict[int, str] = {}
    for i in industrias:
        iid = to_int(i.get("cont_id") or i.get("forn_id") or i.get("ind_id"))
        if iid is None:
            continue
        nome = (i.get("cont_nome_fantasia") or i.get("cont_razao_social")
                or i.get("forn_nome") or i.get("ind_nome") or "").strip()
        if nome:
            mapa[iid] = nome
    print(f"  {len(mapa)} indústrias mapeadas")
    return mapa


def construir_mapa_produtos(host: str, headers: dict) -> dict[int, dict]:
    """
    Tenta paginação primeiro. Se a paginação tiver bug (todas as páginas retornando
    o mesmo conjunto), o mapa vai ficar pequeno — chamadas individuais são feitas
    sob demanda em complementar_mapas() depois que sabemos quais prod_ids são usados.
    """
    print("\n[3/3] Baixando produtos (paginação) ...")
    produtos = fetch_all(host, "/Produto", headers, label="/Produto")
    mapa: dict[int, dict] = {}
    for p in produtos:
        pid = to_int(p.get("prod_id"))
        if pid is None:
            continue
        codigo = (p.get("prod_codigo") or "").strip()
        nome = (p.get("prod_descricao") or p.get("prod_nome") or "").strip()
        mapa[pid] = {"codigo": codigo, "nome": nome}
    print(f"  {len(mapa)} produtos únicos via paginação")
    return mapa


def complementar_mapas(
    host: str,
    headers: dict,
    pedidos: list,
    cliente_nome: dict,
    produto_info: dict,
    ano_alvo: int,
) -> tuple[int, int]:
    """
    Para os pedidos do ano-alvo com faturamento parcial, identifica cli_ids e
    prod_ids ausentes nos mapas e baixa individualmente via /Cliente/{id} e
    /Produto/{id}. Necessário porque a paginação de /Cliente e /Produto tem
    bug que retorna sempre o mesmo conjunto de 100 registros.
    """
    cli_ids_faltantes: set[int] = set()
    prod_ids_faltantes: set[int] = set()

    for p in pedidos:
        if p.get("excluido") == 1:
            continue
        d_fat = parse_iso(p.get("pedi_data_fatura"))
        d_cad = parse_iso(p.get("pedi_data_cadastro"))
        d_ref = d_fat or d_cad
        if d_ref is None or d_ref.year != ano_alvo:
            continue

        itens = p.get("pedi_itens") or []
        algum_fat = any(to_float(it.get("peit_qtde_faturada")) > 0 and it.get("excluido") != 1 for it in itens)
        algum_pend = any(
            to_float(it.get("peit_qtde")) > to_float(it.get("peit_qtde_faturada"))
            and it.get("excluido") != 1
            for it in itens
        )
        if not (algum_fat and algum_pend):
            continue

        cli_id = to_int(p.get("pedi_cont_id"))
        if cli_id and cli_id not in cliente_nome:
            cli_ids_faltantes.add(cli_id)
        for it in itens:
            if it.get("excluido") == 1:
                continue
            if to_float(it.get("peit_qtde")) > to_float(it.get("peit_qtde_faturada")):
                prod_id = to_int(it.get("peit_prod_id"))
                if prod_id and prod_id not in produto_info:
                    prod_ids_faltantes.add(prod_id)

    novos_cli = 0
    if cli_ids_faltantes:
        print(f"\n  complementando {len(cli_ids_faltantes)} clientes via /Cliente/{{id}} ...")
        for cid in cli_ids_faltantes:
            c = fetch_individual(host, headers, "/Cliente", cid)
            if c:
                nome = (c.get("cont_nome_fantasia") or c.get("cont_razao_social") or "").strip()
                if nome:
                    cliente_nome[cid] = nome
                    novos_cli += 1
        print(f"  +{novos_cli} clientes resolvidos")

    novos_prod = 0
    if prod_ids_faltantes:
        print(f"\n  complementando {len(prod_ids_faltantes)} produtos via /Produto/{{id}} ...")
        for i, pid in enumerate(sorted(prod_ids_faltantes), 1):
            p = fetch_individual(host, headers, "/Produto", pid)
            if p:
                codigo = (p.get("prod_codigo") or "").strip()
                nome = (p.get("prod_descricao") or p.get("prod_nome") or "").strip()
                if nome or codigo:
                    produto_info[pid] = {"codigo": codigo, "nome": nome}
                    novos_prod += 1
            if i % 50 == 0:
                print(f"    {i}/{len(prod_ids_faltantes)} ({novos_prod} resolvidos)")
        print(f"  +{novos_prod} produtos resolvidos")

    return novos_cli, novos_prod


def filtrar_e_extrair(
    pedidos: list,
    cliente_nome: dict,
    industria_nome: dict,
    produto_info: dict,
    ano_alvo: int,
    host: str = "",
    headers: dict | None = None,
) -> dict:
    """
    Filtra os pedidos do ano-alvo com faturamento parcial e
    monta a lista de itens com saldo.
    """
    itens_saldo: list[dict] = []
    status_counter: dict[str, int] = defaultdict(int)
    pedidos_2026 = 0
    pedidos_parciais = 0

    # Classificação adicional de TODOS os pedidos 2026
    n_totalmente_faturados = 0
    n_sem_fatura = 0
    n_pendentes_entrega = 0
    n_entregues = 0
    valor_pedidos_faturados = 0.0
    valor_pedidos_total = 0.0  # soma dos valores de TODOS os pedidos 2026
    pedidos_2026_lista: list[dict] = []  # lista resumida para o painel

    for p in pedidos:
        pid = p.get("pedi_id")
        if p.get("excluido") == 1:
            continue
        # Filtro de ano: considera pedi_data_fatura ou pedi_data_cadastro
        d_fat = parse_iso(p.get("pedi_data_fatura"))
        d_cad = parse_iso(p.get("pedi_data_cadastro"))
        d_ref = d_fat or d_cad
        if d_ref is None or d_ref.year != ano_alvo:
            continue
        pedidos_2026 += 1

        # FIX 2026-05-04 v3: bulk /Pedido (sem filtros) já traz pedi_itens correto.
        # Só busca individual se a lista bulk realmente não trouxe (raro com /Pedido sem filtros).
        itens = p.get("pedi_itens") or []
        if not itens and p.get("pedi_id") and host and headers:
            try:
                detalhe = fetch_pedido_individual(host, headers, p["pedi_id"])
                if detalhe and detalhe.get("pedi_itens"):
                    itens = detalhe["pedi_itens"]
            except Exception:
                pass
        # Avalia condição de "faturado parcial" e classifica geral:
        algum_faturado = False
        algum_pendente = False
        itens_validos = []
        valor_total_pedido = 0.0
        for it in itens:
            if it.get("excluido") == 1:
                continue
            qtd = to_float(it.get("peit_qtde"))
            qfat = to_float(it.get("peit_qtde_faturada"))
            if qtd <= 0:
                continue
            if qfat > 0:
                algum_faturado = True
            if qtd > qfat:
                algum_pendente = True
                itens_validos.append((it, qtd, qfat))
            valor_total_pedido += qtd * (to_float(it.get("peit_preco_real")) or to_float(it.get("peit_preco")))

        # Classificação ampla (todos pedidos 2026)
        valor_pedidos_total += valor_total_pedido
        status_envio = (p.get("pedi_status_envio") or "").upper().strip()
        # status_envio: 'N' = não enviado/pendente, 'S' = enviado/entregue
        if status_envio == "S":
            n_entregues += 1
        else:
            # se tem fatura, está pendente de entrega; senão não conta
            if algum_faturado:
                n_pendentes_entrega += 1

        cli_id_geral = to_int(p.get("pedi_cont_id"))
        ind_id_geral = to_int(p.get("pedi_forn_id"))
        cli_nome_geral = cliente_nome.get(cli_id_geral, "") if cli_id_geral else ""
        if not cli_nome_geral:
            co = p.get("cliente") or {}
            cli_nome_geral = (co.get("cont_nome_fantasia") or co.get("cont_razao_social") or "").strip() or (f"Cliente #{cli_id_geral}" if cli_id_geral else "Cliente s/ ID")
        ind_nome_geral = industria_nome.get(ind_id_geral, "") if ind_id_geral else ""
        if not ind_nome_geral and ind_id_geral:
            ind_nome_geral = f"Indústria #{ind_id_geral}"

        if algum_faturado and not algum_pendente:
            n_totalmente_faturados += 1
            valor_pedidos_faturados += valor_total_pedido
        elif algum_faturado and algum_pendente:
            valor_pedidos_faturados += valor_total_pedido  # parcial conta no faturado também
        elif not algum_faturado:
            n_sem_fatura += 1

        # Resumo deste pedido pra usar no painel (Visão Geral)
        pedidos_2026_lista.append({
            "pedido_id": to_int(p.get("pedi_id")),
            "ordem_compra": (p.get("pedi_ordem_compra") or "").strip(),
            "cliente": cli_nome_geral,
            "industria": ind_nome_geral,
            "data": (d_fat or d_cad).date().isoformat() if (d_fat or d_cad) else None,
            "mes": d_ref.month,
            "valor_total": round(valor_total_pedido, 2),
            "fatura_status": "Total" if (algum_faturado and not algum_pendente) else ("Parcial" if (algum_faturado and algum_pendente) else "Sem fatura"),
            "entrega_status": "Entregue" if status_envio == "S" else ("Pendente" if algum_faturado else "Aguardando faturamento"),
            "pedi_status": p.get("pedi_status"),
        })

        if not (algum_faturado and algum_pendente):
            continue
        pedidos_parciais += 1
        status_counter[str(p.get("pedi_status"))] += 1

        cli_id = to_int(p.get("pedi_cont_id"))
        cli_nome = cliente_nome.get(cli_id, "") if cli_id else ""
        if not cli_nome:
            cli_obj = p.get("cliente") or {}
            cli_nome = (cli_obj.get("cont_nome_fantasia") or cli_obj.get("cont_razao_social") or "").strip()
        if not cli_nome:
            cli_nome = (p.get("pedi_cont_nome_fantasia") or p.get("pedi_cont_razao_social")
                        or p.get("pedi_cont_nome") or "").strip()
        if not cli_nome:
            cli_nome = f"Cliente #{cli_id}" if cli_id else "Cliente s/ ID"

        ind_id = to_int(p.get("pedi_forn_id"))
        ind_nome = industria_nome.get(ind_id, "") if ind_id else ""
        if not ind_nome and ind_id:
            ind_nome = f"Indústria #{ind_id}"

        mes_ref = (d_ref.month, d_ref.year)

        # Inferência de status do saldo a partir de pedi_obs (Lais/Rafaela escrevem
        # manualmente quando o saldo será cancelado em vez de faturado).
        obs = (p.get("pedi_obs") or "").strip()
        obs_low = obs.lower()
        cancel_keywords = ("cancel", "retirar saldo", "retira saldo", "anular saldo", "estornar", "descontinu")
        if any(k in obs_low for k in cancel_keywords):
            status_saldo = "Cancelar"
        else:
            status_saldo = "Faturar"

        for it, qtd, qfat in itens_validos:
            saldo = qtd - qfat
            if saldo <= 0:
                continue
            prod_id = to_int(it.get("peit_prod_id"))
            prod = produto_info.get(prod_id) if prod_id else None
            prod_codigo = (prod or {}).get("codigo") or (str(prod_id) if prod_id else "")
            prod_nome = (prod or {}).get("nome") or f"Produto #{prod_id}" if prod_id else "Produto s/ ID"

            preco = to_float(it.get("peit_preco_real")) or to_float(it.get("peit_preco"))
            valor_saldo = round(saldo * preco, 2)

            # Quem atualizou (digitador do pedido) + data atualização do item
            digitador = ""
            dig_obj = p.get("digitador") or {}
            if isinstance(dig_obj, dict):
                digitador = (dig_obj.get("cola_nome") or "").strip()
            data_atualizacao_item = (it.get("ys_datahora_atualizacao") or it.get("ys_datahora") or "")[:10]
            data_atualizacao_pedido = (p.get("ys_datahora_atualizacao") or p.get("ys_datahora") or "")[:10]

            itens_saldo.append({
                "pedido_id": to_int(p.get("pedi_id")),
                "ordem_compra": (p.get("pedi_ordem_compra") or "").strip(),
                "pedi_status": p.get("pedi_status"),
                "status_saldo": status_saldo,
                "observacao": obs,
                "data_fatura": d_fat.date().isoformat() if d_fat else None,
                "data_cadastro": d_cad.date().isoformat() if d_cad else None,
                "data_envio": (p.get("pedi_data_envio") or "")[:10] or None,
                "data_entrega": (p.get("pedi_data_entrega") or "")[:10] or None,
                "mes": mes_ref[0],
                "mes_nome": MESES_PT[mes_ref[0]],
                "ano": mes_ref[1],
                "cliente_id": cli_id,
                "cliente": cli_nome,
                "industria_id": ind_id,
                "industria": ind_nome,
                "produto_id": prod_id,
                "produto_codigo": prod_codigo,
                "produto": prod_nome,
                "qtde_pedida": qtd,
                "qtde_faturada": qfat,
                "qtde_saldo": saldo,
                "preco_unit": preco,
                "valor_saldo": valor_saldo,
                "embalagem": it.get("peit_embalagem", ""),
                "digitador": digitador,
                "data_atualizacao_item": data_atualizacao_item,
                "data_atualizacao_pedido": data_atualizacao_pedido,
            })

    print(f"\n  pedidos do ano {ano_alvo}: {pedidos_2026}")
    print(f"  pedidos parcialmente faturados: {pedidos_parciais}")
    print(f"  itens com saldo a faturar: {len(itens_saldo)}")
    print(f"  status (pedi_status) distribuídos: {dict(status_counter)}")

    return {
        "ano": ano_alvo,
        "data_referencia": datetime.now().date().isoformat(),
        "totais": {
            "pedidos_no_ano": pedidos_2026,
            "pedidos_parcialmente_faturados": pedidos_parciais,
            "pedidos_totalmente_faturados": n_totalmente_faturados,
            "pedidos_sem_fatura": n_sem_fatura,
            "pedidos_pendentes_entrega": n_pendentes_entrega,
            "pedidos_entregues": n_entregues,
            "itens_em_saldo": len(itens_saldo),
            "valor_total_saldo": round(sum(i["valor_saldo"] for i in itens_saldo), 2),
            "valor_pedidos_faturados": round(valor_pedidos_faturados, 2),
            "valor_pedidos_total": round(valor_pedidos_total, 2),
        },
        "status_counter": dict(status_counter),
        "itens": itens_saldo,
        "pedidos_2026": pedidos_2026_lista,
    }


# ============================================================
# AGRUPAMENTOS (cliente / mês / produto)
# ============================================================

def agrupar(itens: list[dict]) -> dict:
    por_cliente: dict[str, dict] = defaultdict(lambda: {"itens": [], "valor": 0.0, "qtd_itens": 0})
    por_mes: dict[str, dict] = defaultdict(lambda: {"itens": [], "valor": 0.0, "qtd_itens": 0})
    por_produto: dict[str, dict] = defaultdict(lambda: {"itens": [], "valor": 0.0, "qtd_itens": 0})
    por_cliente_mes_produto: dict[tuple, dict] = defaultdict(
        lambda: {"itens": [], "valor": 0.0, "qtde_saldo": 0.0}
    )

    for it in itens:
        cli = it["cliente"]
        mes_label = f"{it['mes']:02d}/{it['ano']} ({it['mes_nome']})"
        prod_label = f"{it['produto_codigo']} — {it['produto']}".strip(" —")

        por_cliente[cli]["itens"].append(it)
        por_cliente[cli]["valor"] += it["valor_saldo"]
        por_cliente[cli]["qtd_itens"] += 1

        por_mes[mes_label]["itens"].append(it)
        por_mes[mes_label]["valor"] += it["valor_saldo"]
        por_mes[mes_label]["qtd_itens"] += 1

        por_produto[prod_label]["itens"].append(it)
        por_produto[prod_label]["valor"] += it["valor_saldo"]
        por_produto[prod_label]["qtd_itens"] += 1

        chave = (cli, mes_label, prod_label)
        por_cliente_mes_produto[chave]["itens"].append(it)
        por_cliente_mes_produto[chave]["valor"] += it["valor_saldo"]
        por_cliente_mes_produto[chave]["qtde_saldo"] += it["qtde_saldo"]

    # serialize para JSON
    return {
        "por_cliente": {k: {"valor": round(v["valor"], 2),
                            "qtd_itens": v["qtd_itens"]}
                        for k, v in sorted(por_cliente.items(),
                                           key=lambda x: -x[1]["valor"])},
        "por_mes": {k: {"valor": round(v["valor"], 2),
                        "qtd_itens": v["qtd_itens"]}
                    for k, v in sorted(por_mes.items())},
        "por_produto": {k: {"valor": round(v["valor"], 2),
                            "qtd_itens": v["qtd_itens"]}
                        for k, v in sorted(por_produto.items(),
                                           key=lambda x: -x[1]["valor"])},
        "por_cliente_mes_produto": [
            {
                "cliente": k[0],
                "mes": k[1],
                "produto": k[2],
                "qtde_saldo": round(v["qtde_saldo"], 2),
                "valor_saldo": round(v["valor"], 2),
            }
            for k, v in sorted(por_cliente_mes_produto.items(),
                               key=lambda x: (x[0][0], x[0][1], -x[1]["valor"]))
        ],
    }


# ============================================================
# OUTPUT — Excel
# ============================================================

def gerar_excel(itens: list[dict], grupos: dict, totais: dict, path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERRO: openpyxl não instalado. Rode: pip3 install openpyxl --break-system-packages")
        raise

    wb = Workbook()

    # ---------- Aba 1: Resumo ----------
    ws = wb.active
    ws.title = "Resumo"
    fill_titulo = PatternFill("solid", fgColor="305496")
    fill_hdr = PatternFill("solid", fgColor="8EA9DB")
    fill_zebra = PatternFill("solid", fgColor="F2F2F2")
    font_titulo = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    font_hdr = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_norm = Font(name="Arial", size=10)
    bold = Font(name="Arial", size=10, bold=True)
    border = Border(left=Side(style="thin", color="BFBFBF"),
                    right=Side(style="thin", color="BFBFBF"),
                    top=Side(style="thin", color="BFBFBF"),
                    bottom=Side(style="thin", color="BFBFBF"))

    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value=f"SALDOS DE PEDIDOS — {totais.get('data_referencia', '')}")
    c.font = font_titulo
    c.fill = fill_titulo
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    linha = 3
    for label, val in [
        ("Pedidos no ano", totais.get("pedidos_no_ano", 0)),
        ("Pedidos parcialmente faturados", totais.get("pedidos_parcialmente_faturados", 0)),
        ("Itens em saldo", totais.get("itens_em_saldo", 0)),
        ("Valor total em saldo (R$)", totais.get("valor_total_saldo", 0.0)),
    ]:
        ws.cell(row=linha, column=1, value=label).font = bold
        cv = ws.cell(row=linha, column=2, value=val)
        if "R$" in label:
            cv.number_format = "R$ #,##0.00"
        cv.font = font_norm
        linha += 1

    # tabela "Top clientes"
    linha += 2
    ws.cell(row=linha, column=1, value="Top clientes (por valor em saldo)").font = bold
    linha += 1
    headers_tc = ["Cliente", "Valor saldo (R$)", "Qtd. itens"]
    for i, h in enumerate(headers_tc, start=1):
        cell = ws.cell(row=linha, column=i, value=h)
        cell.font = font_hdr; cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    linha += 1
    for cli, info in list(grupos["por_cliente"].items())[:10]:
        ws.cell(row=linha, column=1, value=cli).font = font_norm
        v = ws.cell(row=linha, column=2, value=info["valor"])
        v.number_format = "R$ #,##0.00"; v.font = font_norm
        ws.cell(row=linha, column=3, value=info["qtd_itens"]).font = font_norm
        linha += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    # ---------- Botão "ATUALIZAR ERP" no canto superior direito ----------
    try:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.worksheet.hyperlink import Hyperlink

        projeto_dir = Path(path).resolve().parent.parent  # outputs/.. == diretório do projeto
        asset_btn = projeto_dir / "assets" / "btn_atualizar_erp.png"
        rodar_cmd = projeto_dir / "rodar.command"

        if asset_btn.exists():
            # Aumenta as colunas F:I e linhas 1-7 pra acomodar o botão
            for col_letter, w in [("F", 14), ("G", 14), ("H", 14), ("I", 4)]:
                ws.column_dimensions[col_letter].width = w
            for r in range(1, 8):
                ws.row_dimensions[r].height = 22

            # Faixa de células sob o botão recebe hyperlink (todas elas, pra clicar em qualquer canto)
            cmd_url = f"file://{rodar_cmd}"
            for r in range(1, 8):
                for col_letter in ("F", "G", "H"):
                    cel = ws[f"{col_letter}{r}"]
                    cel.hyperlink = Hyperlink(ref=cel.coordinate, target=cmd_url, tooltip="Clique pra atualizar dados via API do ERP")

            # Adiciona um caption clicável bem visível
            ws.merge_cells("F8:H8")
            cap = ws["F8"]
            cap.value = "🔄 Atualizar ERP (clique aqui)"
            cap.font = Font(name="Arial", size=11, bold=True, color="0563C1", underline="single")
            cap.alignment = Alignment(horizontal="center", vertical="center")
            cap.hyperlink = Hyperlink(ref="F8", target=cmd_url, tooltip="Clique pra atualizar dados via API do ERP")

            img = XLImage(str(asset_btn))
            # Ajusta tamanho (default é 360x360 mas queremos ~140x140 no Excel)
            img.width = 150
            img.height = 150
            img.anchor = "F1"
            ws.add_image(img)
    except Exception as e:
        # Se Pillow não estiver disponível ou asset faltar, segue sem o botão
        print(f"  (aviso: botão ATUALIZAR ERP não inserido — {e})")

    # ---------- Aba 2: Saldos por Cliente / Mês / Produto ----------
    ws2 = wb.create_sheet("Por Cliente-Mês-Produto")
    headers2 = ["Cliente", "Mês", "Produto", "Qtde Saldo", "Valor Saldo (R$)"]
    for i, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = font_hdr; cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for r, row in enumerate(grupos["por_cliente_mes_produto"], start=2):
        ws2.cell(row=r, column=1, value=row["cliente"]).font = font_norm
        ws2.cell(row=r, column=2, value=row["mes"]).font = font_norm
        ws2.cell(row=r, column=3, value=row["produto"]).font = font_norm
        cq = ws2.cell(row=r, column=4, value=row["qtde_saldo"])
        cq.number_format = "#,##0.00"; cq.font = font_norm
        cv = ws2.cell(row=r, column=5, value=row["valor_saldo"])
        cv.number_format = "R$ #,##0.00"; cv.font = font_norm
        if r % 2 == 0:
            for col in range(1, 6):
                ws2.cell(row=r, column=col).fill = fill_zebra
    ws2.auto_filter.ref = f"A1:E{max(2, len(grupos['por_cliente_mes_produto']) + 1)}"
    ws2.freeze_panes = "A2"
    larguras2 = [38, 22, 50, 14, 18]
    for i, w in enumerate(larguras2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ---------- Aba 3: Por Indústria (formato de envio) ----------
    ws3 = wb.create_sheet("Por Indústria (envio)")
    headers3 = ["Cliente", "Produto", "Cód. Produto", "Qtd Saldo",
                "Valor Saldo (R$)", "Pedido nº", "OC Cliente",
                "Status", "Observação"]
    fill_faturar = PatternFill("solid", fgColor="DCE6F1")    # azul claro
    fill_cancelar = PatternFill("solid", fgColor="FCE4D6")   # rosa claro
    font_faturar = Font(name="Arial", size=10, bold=True, color="1F4E79")
    font_cancelar = Font(name="Arial", size=10, bold=True, color="C00000")

    # Agrupa itens por indústria → ordena por valor desc dentro do grupo
    itens_por_ind: dict[str, list] = defaultdict(list)
    for it in itens:
        itens_por_ind[it.get("industria") or "(sem indústria)"].append(it)
    # Ordena indústrias pelo valor total desc
    inds_ordenadas = sorted(
        itens_por_ind.keys(),
        key=lambda k: -sum(i["valor_saldo"] for i in itens_por_ind[k]),
    )

    fill_ind = PatternFill("solid", fgColor="FFD966")  # amarelo destaque
    font_ind = Font(name="Arial", size=12, bold=True, color="000000")
    fill_subtotal = PatternFill("solid", fgColor="FFF2CC")
    font_subtotal = Font(name="Arial", size=10, bold=True, italic=True)

    r = 1
    for ind_nome in inds_ordenadas:
        bloco = sorted(itens_por_ind[ind_nome], key=lambda x: (-x["valor_saldo"], x["cliente"]))
        valor_grp = sum(i["valor_saldo"] for i in bloco)
        qtd_grp = len(bloco)

        # Cabeçalho da indústria (legenda)
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        c = ws3.cell(row=r, column=1,
                     value=f"INDÚSTRIA: {ind_nome}   —   {qtd_grp} itens   —   R$ {valor_grp:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        c.font = font_ind; c.fill = fill_ind
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws3.row_dimensions[r].height = 22
        r += 1

        # Cabeçalho de colunas
        for i, h in enumerate(headers3, start=1):
            cell = ws3.cell(row=r, column=i, value=h)
            cell.font = font_hdr; cell.fill = fill_hdr
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        r += 1

        # Linhas de itens
        for it in bloco:
            ws3.cell(row=r, column=1, value=it["cliente"]).font = font_norm
            ws3.cell(row=r, column=2, value=it["produto"]).font = font_norm
            ws3.cell(row=r, column=3, value=it.get("produto_codigo", "")).font = font_norm
            cq = ws3.cell(row=r, column=4, value=it["qtde_saldo"])
            cq.number_format = "#,##0.00"; cq.font = font_norm
            cv = ws3.cell(row=r, column=5, value=it["valor_saldo"])
            cv.number_format = "R$ #,##0.00"; cv.font = font_norm
            ws3.cell(row=r, column=6, value=it["pedido_id"]).font = font_norm
            ws3.cell(row=r, column=7, value=it.get("ordem_compra", "")).font = font_norm
            # Status (Faturar / Cancelar) com cor condicional
            cell_status = ws3.cell(row=r, column=8, value=it.get("status_saldo", "Faturar"))
            if it.get("status_saldo") == "Cancelar":
                cell_status.font = font_cancelar; cell_status.fill = fill_cancelar
            else:
                cell_status.font = font_faturar; cell_status.fill = fill_faturar
            cell_status.alignment = Alignment(horizontal="center")
            # Observação (pedi_obs)
            ws3.cell(row=r, column=9, value=it.get("observacao", "")).font = font_norm
            r += 1

        # Subtotal da indústria
        ws3.cell(row=r, column=1, value=f"Subtotal — {ind_nome}").font = font_subtotal
        ws3.cell(row=r, column=1).fill = fill_subtotal
        for col in range(2, 4):
            ws3.cell(row=r, column=col).fill = fill_subtotal
        cqs = ws3.cell(row=r, column=4, value=sum(i["qtde_saldo"] for i in bloco))
        cqs.number_format = "#,##0.00"; cqs.font = font_subtotal; cqs.fill = fill_subtotal
        cvs = ws3.cell(row=r, column=5, value=valor_grp)
        cvs.number_format = "R$ #,##0.00"; cvs.font = font_subtotal; cvs.fill = fill_subtotal
        for col in (6, 7, 8, 9):
            ws3.cell(row=r, column=col).fill = fill_subtotal
        r += 2  # linha em branco entre indústrias

    ws3.freeze_panes = "A2"
    larguras3 = [34, 50, 14, 12, 16, 11, 16, 12, 40]
    for i, w in enumerate(larguras3, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ---------- Aba 4: Detalhe item-a-item ----------
    ws4 = wb.create_sheet("Itens (detalhado)")
    headers4 = ["Pedido", "OC Cliente", "Status Saldo", "Dt. Fatura", "Mês", "Cliente", "Indústria",
                "Cód. Produto", "Produto", "Embalagem",
                "Qtd Pedida", "Qtd Faturada", "Qtd Saldo",
                "Preço Unit. (R$)", "Valor Saldo (R$)", "Observação"]
    for i, h in enumerate(headers4, start=1):
        cell = ws4.cell(row=1, column=i, value=h)
        cell.font = font_hdr; cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for r, it in enumerate(itens, start=2):
        vals = [
            it["pedido_id"], it.get("ordem_compra", ""),
            it.get("status_saldo", "Faturar"), it.get("data_fatura"),
            f"{it['mes']:02d}/{it['ano']}",
            it["cliente"], it["industria"],
            it["produto_codigo"], it["produto"], it.get("embalagem", ""),
            it["qtde_pedida"], it["qtde_faturada"], it["qtde_saldo"],
            it["preco_unit"], it["valor_saldo"], it.get("observacao", ""),
        ]
        for c_idx, v in enumerate(vals, start=1):
            cell = ws4.cell(row=r, column=c_idx, value=v)
            cell.font = font_norm
        # Cor condicional na coluna Status Saldo (col 3)
        cell_status = ws4.cell(row=r, column=3)
        if it.get("status_saldo") == "Cancelar":
            cell_status.font = font_cancelar; cell_status.fill = fill_cancelar
        else:
            cell_status.font = font_faturar; cell_status.fill = fill_faturar
        cell_status.alignment = Alignment(horizontal="center")
        ws4.cell(row=r, column=4).number_format = "DD/MM/YYYY"
        ws4.cell(row=r, column=11).number_format = "#,##0.00"
        ws4.cell(row=r, column=12).number_format = "#,##0.00"
        ws4.cell(row=r, column=13).number_format = "#,##0.00"
        ws4.cell(row=r, column=14).number_format = "R$ #,##0.0000"
        ws4.cell(row=r, column=15).number_format = "R$ #,##0.00"
        if r % 2 == 0:
            for col in range(1, 17):
                # Não pinta a coluna Status com zebra (cor já é condicional)
                if col == 3:
                    continue
                ws4.cell(row=r, column=col).fill = fill_zebra
    ws4.auto_filter.ref = f"A1:P{max(2, len(itens) + 1)}"
    ws4.freeze_panes = "A2"
    larguras4 = [9, 16, 12, 12, 10, 32, 22, 12, 38, 11, 11, 12, 11, 14, 16, 40]
    for i, w in enumerate(larguras4, start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    print(f"\n  Excel salvo em: {path}")


# ============================================================
# MAIN
# ============================================================

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    parser.add_argument("--config", required=True, help="Caminho do config.json da skill")
    parser.add_argument("--output-json", required=True, help="Caminho do JSON de saída")
    parser.add_argument("--output-xlsx", required=True, help="Caminho do Excel de saída")
    parser.add_argument("--ano", type=int, default=None, help="Ano-alvo (default vem do config.json ou 2026)")
    args = parser.parse_args(argv)

    cfg_path = Path(args.config)
    if not cfg_path.exists():
        print(f"ERRO: config não encontrada em {cfg_path}", file=sys.stderr)
        return 2
    cfg = json.loads(cfg_path.read_text(encoding="utf-8"))

    drive_root = find_drive_root()
    creds_rel = cfg.get("credenciais_relativa", "3MV/Automações/comissoes/credenciais.json")
    creds_path = drive_root / creds_rel
    creds = carregar_credenciais(creds_path)

    host = cfg.get("api_host", DEFAULT_HOST)
    headers = montar_headers(creds)
    ano_alvo = args.ano or cfg.get("ano_alvo", 2026)

    print(f"=== EXTRAIR SALDOS — ano {ano_alvo} ===")
    print(f"  host: {host}")
    print(f"  cliente API: {headers.get('Cliente')}")

    cliente_nome = construir_mapa_clientes(host, headers)
    industria_nome = construir_mapa_industrias(host, headers)
    produto_info = construir_mapa_produtos(host, headers)

    print("\n[Pedidos] paginação híbrida ...")
    pedidos = fetch_pedidos_completo(host, headers)

    # Complementa mapas de cliente e produto via GET individual (necessário
    # porque /Cliente e /Produto têm bug de paginação que retorna sempre os
    # mesmos 100 registros, deixando muitos IDs sem nome)
    complementar_mapas(host, headers, pedidos, cliente_nome, produto_info, ano_alvo)

    resultado = filtrar_e_extrair(pedidos, cliente_nome, industria_nome, produto_info, ano_alvo, host=host, headers=headers)
    grupos = agrupar(resultado["itens"])
    resultado["agrupamentos"] = grupos

    out_json = Path(args.output_json)
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(resultado, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n  JSON salvo em: {out_json}")

    out_xlsx = Path(args.output_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    gerar_excel(resultado["itens"], grupos, resultado["totais"], out_xlsx)

    return 0


if __name__ == "__main__":
    sys.exit(main())
