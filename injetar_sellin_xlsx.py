#!/usr/bin/env python3
"""
injetar_sellin_xlsx.py
Roda APÓS injetar_categorias_precos.py.

Em vez de bater na API (lenta — leva 5-15 min), lê direto os Excel:
  - planilhas-snapshot/sellout_2025.xlsx
  - planilhas-snapshot/sellout_2026.xlsx
  (e opcionalmente sellout_2024.xlsx se existir)

Cruza CODPROD/CODFORNEC com mapa_produtos_api (já injetado) pra pegar categoria.
Quando não encontra categoria na API, usa fallback Excel da pasta:
  /3MV/R/SUAS VENDAS/{INDUSTRIA}/*.xlsx
  (procura por CÓDIGO ou nome no Excel da indústria; categoria = nome da pasta)

Injeta em dados.json:
  - sellin_historico = {industria: {categoria: {2024: [m1..m12], 2025: [...], 2026: [...]}}}
  - sellin_cli_ind_cat = {cliente: {industria: {categoria: {ano: [meses]}}}}

Caminho dos xlsx: relativo ao próprio script (planilhas-snapshot/) + Drive.
"""
from __future__ import annotations
import json
import sys
import os
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERRO: openpyxl não instalado. Rode: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


HERE = Path(__file__).parent
# DRIVE descoberto dinamicamente: tenta múltiplos caminhos (Mac, GH Actions, sandbox)
_drive_candidatos = [
    Path.home() / "Library/CloudStorage/GoogleDrive-marcelo@3mvrepresentacao.com/Meu Drive/3MV",
    Path("/Users/marcelorodrigues/Library/CloudStorage/GoogleDrive-marcelo@3mvrepresentacao.com/Meu Drive/3MV"),
    # Sandbox mounts (Claude/Cowork)
    Path("/sessions/festive-zealous-faraday/mnt/3MV"),
    Path("/sessions/pensive-vibrant-sagan/mnt/GoogleDrive-marcelo@3mvrepresentacao.com/Meu Drive/3MV"),
    # Subindo da pasta do script (~/3MV/Automações/painel-3mv-cron → ~/3MV)
    HERE.parent.parent,
]
DRIVE = next((p for p in _drive_candidatos if p.exists()), HERE.parent.parent)
SNAPSHOT_DIR = HERE / "planilhas-snapshot"

# Caminhos prováveis pros sellout xlsx
SELLOUT_CANDIDATOS = [
    SNAPSHOT_DIR / "sellout_{ano}.xlsx",
    DRIVE / "BI/Sell-out/sellout_{ano}.xlsx",
    DRIVE / "BI/sellout_{ano}.xlsx",
    HERE / "sellout_{ano}.xlsx",
]

# Pasta com categorias por indústria (fallback quando API não tem)
# Tenta caminho exato "R/SUAS VENDAS" e variações
PASTA_SUAS_VENDAS = None
for sufx in ("R/SUAS VENDAS", "R/Suas Vendas", "r/suas vendas", "SUAS VENDAS", "Suas Vendas"):
    p = DRIVE / sufx
    if p.exists():
        PASTA_SUAS_VENDAS = p
        break
if PASTA_SUAS_VENDAS is None:
    PASTA_SUAS_VENDAS = DRIVE / "R/SUAS VENDAS"  # placeholder; vai dar miss se não existir


def to_int(v):
    if v is None: return None
    try: return int(str(v).strip())
    except Exception: return None


def to_float(v):
    if v is None: return 0.0
    try: return float(str(v).replace(",", ".").strip())
    except Exception: return 0.0


def to_str(v):
    if v is None: return ""
    return str(v).strip()


def parse_data(v):
    """Aceita datetime ou string ISO/BR."""
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    s = str(v).strip()
    if not s: return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S"):
        try: return datetime.strptime(s[:19], fmt).date()
        except Exception: pass
    try: return datetime.fromisoformat(s).date()
    except Exception: return None


def carregar_fallback_excel_categorias():
    """Lê /R/SUAS VENDAS/{INDUSTRIA}/*.xlsx → constrói mapa codigo → (industria, categoria)."""
    mapa = {}  # codigo → {industria, categoria, preco, nome_produto}
    if not PASTA_SUAS_VENDAS.exists():
        print(f"  ⚠ Pasta {PASTA_SUAS_VENDAS} não existe — pulando fallback Excel")
        return mapa

    indústrias_pasta = sorted([p for p in PASTA_SUAS_VENDAS.iterdir() if p.is_dir()])
    print(f"  Encontradas {len(indústrias_pasta)} pastas de indústria")
    for ind_dir in indústrias_pasta:
        industria_nome = ind_dir.name.strip().upper()
        # Pega o Excel mais recente da pasta
        xlsxs = sorted(ind_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not xlsxs:
            continue
        x = xlsxs[0]
        try:
            wb = load_workbook(x, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header: continue
            # Tenta achar colunas comuns: Código, Nome do produto, Preço
            col_cod = col_nome = col_preco = None
            for i, h in enumerate(header):
                if h is None: continue
                hn = str(h).strip().lower()
                if col_cod is None and ("código" in hn or "codigo" in hn or hn == "cod"):
                    col_cod = i
                elif col_nome is None and ("nome" in hn or "descricao" in hn or "descrição" in hn or "produto" in hn):
                    col_nome = i
                elif col_preco is None and "preço" in hn or "preco" in hn:
                    col_preco = i
            if col_cod is None:
                col_cod = 0  # default: primeira coluna
            count = 0
            for r in rows:
                if not r or len(r) <= col_cod: continue
                cod = to_str(r[col_cod])
                if not cod: continue
                nome = to_str(r[col_nome]) if col_nome is not None and col_nome < len(r) else ""
                preco = to_float(r[col_preco]) if col_preco is not None and col_preco < len(r) else 0.0
                if cod not in mapa:  # mantém só primeira ocorrência por código
                    mapa[cod] = {
                        "industria": industria_nome,
                        "categoria": industria_nome + " · OUTROS",  # subcategoria genérica
                        "preco": preco,
                        "nome_produto": nome,
                        "fonte": "excel_fallback"
                    }
                    count += 1
            wb.close()
            print(f"    ✓ {industria_nome}: {count} produtos lidos de {x.name}")
        except Exception as e:
            print(f"    ✗ {industria_nome}: erro lendo {x.name} — {e}")
    print(f"  Total fallback Excel: {len(mapa)} produtos")
    return mapa


def carregar_sellout_xlsx(ano: int):
    """Procura sellout_{ano}.xlsx e retorna lista de linhas."""
    for tmpl in SELLOUT_CANDIDATOS:
        path = Path(str(tmpl).replace("{ano}", str(ano)))
        if path.exists():
            print(f"  Lendo {path.name} ({path.stat().st_size/1024:.0f} KB)...")
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter)
            # Indice das colunas
            col_idx = {}
            for i, h in enumerate(header):
                if h is None: continue
                col_idx[str(h).strip()] = i

            linhas = []
            for r in rows_iter:
                if not r: continue
                def G(name):
                    i = col_idx.get(name)
                    return r[i] if i is not None and i < len(r) else None
                linhas.append({
                    "data": parse_data(G("DATA") or G("DTFAT")),
                    "cliente": to_str(G("CLIENTE") or G("FANTASIA")),
                    "cod_prod": to_str(G("CODPROD") or G("CODPRODPRINC")),
                    "descricao": to_str(G("DESCRICAO")),
                    "fornecedor": to_str(G("FORNECEDOR") or G("FORNECDOR PRINCIPAL") or G("FORNECEDOR PRINCIPAL")),
                    "qt": to_float(G("QT")),
                    "vlvenda": to_float(G("VLVENDA")),
                    "ramo": to_str(G("RAMO")),
                    "uf": to_str(G("UF")),
                })
            wb.close()
            print(f"    ✓ {len(linhas)} linhas lidas")
            return linhas
    print(f"  ⚠ Sellout {ano} não encontrado nos caminhos: " + ", ".join(str(c).replace("{ano}", str(ano)) for c in SELLOUT_CANDIDATOS))
    return []


def main():
    dados_path = HERE / "dados.json"
    if not dados_path.exists():
        print(f"ERRO: dados.json não encontrado em {dados_path}", file=sys.stderr)
        return 1
    d = json.loads(dados_path.read_text(encoding="utf-8"))

    print("→ injetar_sellin_xlsx.py")
    print(f"  DRIVE detectado: {DRIVE} (exists={DRIVE.exists()})")
    print(f"  PASTA_SUAS_VENDAS: {PASTA_SUAS_VENDAS} (exists={PASTA_SUAS_VENDAS.exists()})")
    print(f"  dados.json: {dados_path}  ({dados_path.stat().st_size/1024:.0f} KB)")

    # Mapa de categoria existente (API)
    mapa_api = d.get("mapa_produtos_api", {}) or {}
    print(f"  Mapa produtos API: {len(mapa_api)} entradas")

    # Carrega fallback Excel
    print(f"\n→ Fallback Excel (/R/SUAS VENDAS):")
    fallback = carregar_fallback_excel_categorias()

    # Lê sellout 2025 e 2026 (e tenta 2024)
    print(f"\n→ Lendo Sell-out xlsx:")
    sellout_por_ano = {}
    for ano in (2024, 2025, 2026):
        linhas = carregar_sellout_xlsx(ano)
        if linhas:
            sellout_por_ano[ano] = linhas

    if not sellout_por_ano:
        print("ERRO: nenhum sellout encontrado", file=sys.stderr)
        return 1

    # Agrega: sellin_historico = {industria: {categoria: {ano: [m1..m12]}}}
    sellin_historico = {}
    sellin_cli_ind_cat = {}
    com_cat_api = 0
    com_cat_fallback = 0
    sem_cat = 0

    # Mapa de subcategorias por palavras-chave na DESCRICAO (fallback heurístico)
    # Atende JACQUES JANINE, DANA, TIA SONIA, KANITZ, RUBY ROSE etc
    SUBCAT_KEYWORDS = [
        ("MASCARA", ["MASC ", "MASCARA", "MÁSCARA"]),
        ("CONDICIONADOR", ["COND ", "CONDIC"]),
        ("SHAMPOO", ["SHAMP", "SH "]),
        ("LEAVE-IN", ["LEAVE"]),
        ("SPRAY", ["SPRAY"]),
        ("KIT", ["KIT "]),
        ("CREME", ["CREME", "CR "]),
        ("DESODORANTE", ["DESOD", "AERO ANTITR"]),
        ("ROLLON", ["ROLLON", "ROLL ON", "ROLL-ON"]),
        ("TALCO", ["TALCO"]),
        ("SABONETE", ["SABONETE", "SAB "]),
        ("BATOM", ["BATOM"]),
        ("BASE", ["BASE "]),
        ("CHÁ", ["CHA ", "CHÁ"]),
        ("GRANOLA", ["GRANOLA"]),
        ("CASTANHA", ["CASTANHA"]),
        ("CERVEJA", ["CERVEJA"]),
        ("OSSO", ["OSSO"]),
        ("CONDIMENTO", ["TEMPERO", "CONDIM"]),
    ]
    def detectar_subcategoria(descricao: str) -> str:
        if not descricao: return "OUTROS"
        up = descricao.upper()
        for sub, keywords in SUBCAT_KEYWORDS:
            if any(k in up for k in keywords):
                return sub
        return "OUTROS"

    for ano, linhas in sellout_por_ano.items():
        for L in linhas:
            cod = L["cod_prod"]
            ind = L["fornecedor"].upper()
            descricao = L["descricao"]
            # Categoria: API (por código) → fallback Excel (por código) → heurística DESCRICAO
            cat_info = mapa_api.get(cod) if cod else None
            if cat_info and cat_info.get("categoria"):
                cat = cat_info["categoria"]
                if " · " not in cat and ind:
                    cat = ind + " · " + cat  # normaliza pra "INDÚSTRIA · SUBCATEGORIA"
                com_cat_api += 1
            elif cod in fallback:
                cat = fallback[cod]["categoria"]
                if not ind: ind = fallback[cod]["industria"]
                com_cat_fallback += 1
            else:
                # Heurística via DESCRICAO — extrai subcategoria a partir do nome do produto
                sub = detectar_subcategoria(descricao)
                cat = (ind or "OUTROS") + " · " + sub
                sem_cat += 1
            if not ind:
                ind = "DESCONHECIDO"

            dt = L["data"]
            if not dt: continue
            if dt.year != ano: continue  # só conta linhas do ano que estamos processando
            mes = dt.month - 1
            valor = L["vlvenda"]

            # Aggregate industria × categoria × ano × mes
            sellin_historico.setdefault(ind, {}).setdefault(cat, {})
            ano_arr = sellin_historico[ind][cat].setdefault(str(ano), [0.0]*12)
            ano_arr[mes] += valor

            # Aggregate cliente × industria × categoria × ano × mes
            cli = L["cliente"]
            if cli:
                sellin_cli_ind_cat.setdefault(cli, {}).setdefault(ind, {}).setdefault(cat, {})
                cli_arr = sellin_cli_ind_cat[cli][ind][cat].setdefault(str(ano), [0.0]*12)
                cli_arr[mes] += valor

    # Arredonda
    def arred(obj):
        if isinstance(obj, list): return [round(v, 2) for v in obj]
        if isinstance(obj, dict): return {k: arred(v) for k,v in obj.items()}
        return obj
    sellin_historico = arred(sellin_historico)
    sellin_cli_ind_cat = arred(sellin_cli_ind_cat)

    # Conta totais
    total_linhas = sum(len(l) for l in sellout_por_ano.values())
    print(f"\n→ Resumo da injeção:")
    print(f"  Total linhas processadas:     {total_linhas:,}")
    print(f"  Categoria via API:            {com_cat_api:,}")
    print(f"  Categoria via Excel fallback: {com_cat_fallback:,}")
    print(f"  Sem categoria (genérica):     {sem_cat:,}")
    print(f"  Indústrias em sellin_historico: {len(sellin_historico)}")
    print(f"  Clientes em sellin_cli_ind_cat: {len(sellin_cli_ind_cat)}")

    d["sellin_historico"] = sellin_historico
    d["sellin_cli_ind_cat"] = sellin_cli_ind_cat
    d["sellin_historico_meta"] = {
        "fonte": "xlsx (planilhas-snapshot)",
        "total_linhas": total_linhas,
        "anos_processados": sorted(sellout_por_ano.keys()),
        "com_cat_api": com_cat_api,
        "com_cat_excel_fallback": com_cat_fallback,
        "sem_cat": sem_cat,
        "fallback_excel_total": len(fallback),
        "gerado_em": datetime.now().isoformat(),
    }

    dados_path.write_text(json.dumps(d, ensure_ascii=False), encoding="utf-8")
    print(f"\n✓ dados.json regravado ({dados_path.stat().st_size/1024:.0f} KB)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
