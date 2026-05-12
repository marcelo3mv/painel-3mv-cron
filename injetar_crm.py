#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Injeta dados de /3MV/CRM 3MV/CRM_3MV.xlsx no dados.json do painel.

Abas lidas:
  • CRM Indústrias       → contatos por setor de cada indústria (financeiro, adm, mkt, etc)
  • Base de Clientes     → emails, telefones, status, canal/setor de cada cliente
  • Evolução Anual       → valores 2023-2026 por cliente (já temos via API, ignoramos)
  • Contatos Adicionais  → emails extras por indústria/setor

Adiciona ao dados.json:
  d['crm'] = {
    'industrias': { 'BR SPICES': {razao_social, cnpj, status_erp, contatos: {financeiro: {nome, email, tel}, adm: {...}, ...}, lead_time, obs}, ... },
    'clientes':   { 'CLIENTE X': {nome_fantasia, grupo, canal, razao, cnpj, cidade, uf, telefone, email, status, comprou_23_26}, ... },
    'contatos':   { 'CLIENTE X': 'email@...' },     # email principal por cliente — compat
    'setores':    { 'CLIENTE X': 'Açaí' },          # canal vira setor — compat
    'contatos_extras': [{industria, setor, contato, cargo, email, telefone, obs}, ...],
    'planilha_mtime': 'YYYY-MM-DD HH:MM',
    'planilha_path': '/path/to/CRM_3MV.xlsx',
  }
"""
import json
import sys
from pathlib import Path
from datetime import datetime
try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install',
                            '--break-system-packages', '--quiet', 'openpyxl'])
    from openpyxl import load_workbook


def _gd_base():
    candidates = [
        Path.home() / 'Library/CloudStorage/GoogleDrive-marcelo@3mvrepresentacao.com/Meu Drive',
    ]
    for c in candidates:
        if (c / '3MV/CRM 3MV/CRM_3MV.xlsx').exists():
            return c
    return candidates[0]


GD = _gd_base()

# Fallback para GitHub Actions: usar planilhas-snapshot/ do repo se Drive nao existe
_SNAPSHOT = Path(__file__).parent / 'planilhas-snapshot' / 'CRM_3MV.xlsx'
CRM_XLSX = GD / '3MV/CRM 3MV/CRM_3MV.xlsx'
if not CRM_XLSX.exists() and _SNAPSHOT.exists():
    CRM_XLSX = _SNAPSHOT
    print(f'  -> usando snapshot do repo: {CRM_XLSX}')

DADOS_JSON = Path(__file__).parent / 'dados.json'


def _val(v):
    if v is None: return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)):
        return v
    return str(v).strip()


def ler_industrias(wb):
    """CRM Indústrias: ID, Indústria, Razão Social, CNPJ, Status ERP,
       e tripletas Setor — contato/email/tel pra: Financeiro, Contas a Pagar,
       Contas a Receber, Marketing, ADM, Gerência, Supervisão, Trocas, NFD, Encarte
       + Lead time entrega + Observações"""
    sheet = None
    for sn in wb.sheetnames:
        if 'INDUSTRIA' in sn.upper().replace('Ú','U'):
            sheet = wb[sn]; break
    if not sheet:
        return {}
    setores = ['financeiro', 'contas_a_pagar', 'contas_a_receber', 'marketing',
               'adm', 'gerencia', 'supervisao', 'trocas', 'nfd', 'encarte']
    industrias = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]: continue
        nome = _val(row[1]).upper().strip()
        d = {
            'razao_social': _val(row[2]),
            'cnpj': _val(row[3]),
            'status_erp': _val(row[4]),
            'contatos': {},
        }
        # Cada setor ocupa 3 colunas (contato, email, telefone) começando em col 5 (idx 5)
        for i, setor in enumerate(setores):
            base = 5 + i * 3
            if base + 2 < len(row):
                contato = _val(row[base])
                email = _val(row[base+1])
                tel = _val(row[base+2])
                if contato or email or tel:
                    d['contatos'][setor] = {
                        'contato': contato, 'email': email, 'telefone': tel,
                    }
        # Lead time + Observações nos últimos 2
        try:
            d['lead_time_entrega_dias'] = _val(row[35]) if len(row) > 35 else ''
            d['observacoes'] = _val(row[36]) if len(row) > 36 else ''
        except Exception:
            pass
        industrias[nome] = d
    return industrias


def ler_clientes(wb):
    """Base de Clientes: ID, Nome fantasia, Grupo/Rede, Canal, Razão social, CNPJ,
       Cidade, UF, Telefone, E-mail, Pessoa, Status, Comprou 2023-26, ...
       Logística: Agendado, Contato, E-mail, Tel, Tel logística, Resp, Obs"""
    sheet = None
    for sn in wb.sheetnames:
        if 'CLIENTE' in sn.upper() and 'CONTATO' not in sn.upper():
            sheet = wb[sn]; break
    if not sheet:
        return {}, {}, {}
    clientes = {}
    contatos_simples = {}  # nome → email principal
    setores = {}           # nome → canal/setor
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]: continue
        nome = _val(row[1]).upper().strip()
        d = {
            'nome_fantasia': _val(row[1]),
            'grupo': _val(row[2]) if len(row) > 2 else '',
            'canal': _val(row[3]) if len(row) > 3 else '',
            'razao_social': _val(row[4]) if len(row) > 4 else '',
            'cnpj': _val(row[5]) if len(row) > 5 else '',
            'cidade': _val(row[6]) if len(row) > 6 else '',
            'uf': _val(row[7]) if len(row) > 7 else '',
            'telefone': _val(row[8]) if len(row) > 8 else '',
            'email': _val(row[9]) if len(row) > 9 else '',
            'pessoa': _val(row[10]) if len(row) > 10 else '',
            'status': _val(row[11]) if len(row) > 11 else '',
            'comprou_23_26': _val(row[12]) if len(row) > 12 else '',
            'logistica': {
                'agendado': _val(row[13]) if len(row) > 13 else '',
                'contato_agendamento': _val(row[14]) if len(row) > 14 else '',
                'email_agendamento': _val(row[15]) if len(row) > 15 else '',
                'tel_agendamento': _val(row[16]) if len(row) > 16 else '',
                'tel_logistica': _val(row[17]) if len(row) > 17 else '',
                'responsavel': _val(row[18]) if len(row) > 18 else '',
                'observacoes': _val(row[19]) if len(row) > 19 else '',
            },
            # Comprador (colunas U/V/W/X - adicionadas 12/05/2026)
            'central': _val(row[20]) if len(row) > 20 else '',
            'comprador': _val(row[21]) if len(row) > 21 else '',
            'comprador_email': _val(row[22]) if len(row) > 22 else '',
            'comprador_contato': _val(row[23]) if len(row) > 23 else '',
        }
        clientes[nome] = d
        # Compat: campos planos pra usar no painel (contatos + setores)
        if d['email']:
            contatos_simples[nome] = d['email']
        if d['canal']:
            setores[nome] = d['canal']
    return clientes, contatos_simples, setores


def ler_contatos_adicionais(wb):
    """Contatos Adicionais: Indústria, Setor, Contato, Cargo, E-mail, Telefone, Observações"""
    sheet = None
    for sn in wb.sheetnames:
        if 'ADICIONAL' in sn.upper() or 'EXTRA' in sn.upper():
            sheet = wb[sn]; break
    if not sheet:
        return []
    extras = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        d = {
            'industria': _val(row[0]),
            'setor': _val(row[1]) if len(row) > 1 else '',
            'contato': _val(row[2]) if len(row) > 2 else '',
            'cargo': _val(row[3]) if len(row) > 3 else '',
            'email': _val(row[4]) if len(row) > 4 else '',
            'telefone': _val(row[5]) if len(row) > 5 else '',
            'observacoes': _val(row[6]) if len(row) > 6 else '',
        }
        # Só adiciona se tem pelo menos email ou telefone
        if d['email'] or d['telefone']:
            extras.append(d)
    return extras


def main():
    print(f"\n→ Lendo CRM: {CRM_XLSX}")
    if not CRM_XLSX.exists():
        print(f'  ⚠ {CRM_XLSX} não existe — pulando injeção')
        return 0

    wb = load_workbook(CRM_XLSX, data_only=True)
    print(f'  Abas: {wb.sheetnames}')

    industrias = ler_industrias(wb)
    clientes, contatos_simples, setores = ler_clientes(wb)
    extras = ler_contatos_adicionais(wb)

    print(f"  Indústrias com contatos: {len(industrias)}")
    print(f"  Clientes na base: {len(clientes)}")
    print(f"  Clientes com email: {len(contatos_simples)}")
    print(f"  Clientes com canal/setor: {len(setores)}")
    print(f"  Contatos adicionais: {len(extras)}")

    if not DADOS_JSON.exists():
        print(f"  ⚠ dados.json não existe — pulando injeção")
        return 0
    d = json.loads(DADOS_JSON.read_text(encoding='utf-8'))

    try:
        planilha_mtime = datetime.fromtimestamp(CRM_XLSX.stat().st_mtime).strftime('%Y-%m-%d %H:%M')
    except Exception:
        planilha_mtime = ''

    d['crm'] = {
        '_doc': 'Lido de /3MV/CRM 3MV/CRM_3MV.xlsx — atualiza automaticamente a cada execução do pipeline.',
        '_atualizado_em': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'industrias': industrias,
        'clientes': clientes,
        'contatos': contatos_simples,    # compat: usado no painel pra mailto
        'setores': setores,              # compat: usado no painel pra filtro de setor
        'contatos_extras': extras,
        'planilha_mtime': planilha_mtime,
        'planilha_path': str(CRM_XLSX),
    }

    DADOS_JSON.write_text(
        json.dumps(d, ensure_ascii=False),
        encoding='utf-8',
    )
    print(f"✓ dados.json atualizado com seção 'crm'")
    return 0


if __name__ == '__main__':
    sys.exit(main() or 0)
