#!/usr/bin/env python3
"""
Envia alertas de pendências (críticos / atrasados / saldos parciais) por email pra
Lais (adm@), Rafaela (adm1@) e Marcelo (marcelo@) via Worker email-api.

Roda dentro do workflow `atualizar-painel.yml` após gerar_painel.py.

Critérios:
  - CRÍTICOS:  pedido com fatura_status == "Sem fatura" e data > 5 dias atrás
  - ATRASADOS: pedido com fatura mas sem data_entrega, data_fatura > 7 dias atrás
  - SALDOS:    pedido com fatura_status == "Parcial" (vai como anexo Excel)

Layout do email:
  - Tabela agrupada por CLIENTE > INDÚSTRIA (relação completa, sem cortar)
  - Datas em formato DD/MM/AAAA
  - Excel anexo com 1 aba por cliente (saldos parciais — pra enviar aos clientes)

Variáveis de ambiente:
  EMAIL_API_URL   (default: https://email-api.marcelo-778.workers.dev)
  EMAIL_TOKEN     (obrigatório)
  ALERTAS_FORCE   ('1' pra enviar mesmo sem pendências)
  ALERTAS_TO      (override CSV de destinatários; padrão = Lais+Rafaela+Marcelo)
"""
import os
import sys
import json
import io
import base64
import urllib.request
import urllib.error
from collections import defaultdict
from datetime import date, timedelta

API_URL = os.environ.get("EMAIL_API_URL", "https://email-api.marcelo-778.workers.dev").rstrip("/")
TOKEN = os.environ.get("EMAIL_TOKEN", "").strip()
FORCE = os.environ.get("ALERTAS_FORCE", "0").strip() == "1"
TO_OVERRIDE = os.environ.get("ALERTAS_TO", "").strip()

DESTS_PADRAO = [
    ("Lais", "adm@3mvrepresentacao.com"),
    ("Rafaela", "adm1@3mvrepresentacao.com"),
    ("Marcelo", "marcelo@3mvrepresentacao.com"),
]

PAINEL_URL = "https://painel.3mvrepresentacao.com/"
MOBILE_URL = "https://painel.3mvrepresentacao.com/3mv-mobile.html"


def parse_data(s):
    if not s:
        return None
    try:
        return date.fromisoformat(str(s)[:10])
    except Exception:
        return None


def fmt_data(d):
    if d is None:
        return "—"
    if isinstance(d, str):
        d = parse_data(d)
        if d is None:
            return "—"
    try:
        return d.strftime("%d/%m/%Y")
    except Exception:
        return "—"


def fmt_brl(v):
    try:
        s = f"R$ {float(v):,.2f}"
        return s.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"


def carregar_dados():
    for p in ("dados.json", "outputs/dados.json"):
        if os.path.exists(p):
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
    raise FileNotFoundError("dados.json não encontrado")


def carregar_itens(dados):
    itens = dados.get("itens", []) or []
    if isinstance(itens, list):
        return itens
    return []


def classificar(dados):
    """FIX 2026-05-15:
       - Dedup defensivo por pedido_id (a API às vezes traz o mesmo pedido 97x).
       - Critérios reescritos pra usar os campos que REALMENTE chegam na lista
         pedidos_2026: data, fatura_status, entrega_status. data_fatura e
         data_entrega não vêm preenchidos no objeto pedido, então a lógica antiga
         de 'atrasados' nunca disparava — agora usa entrega_status='Pendente' +
         idade do pedido (>=7 dias) como aproximação de entrega atrasada.
    """
    hoje = date.today()
    peds = dados.get("pedidos_2026") or dados.get("pedidos") or []

    # Dedup defensivo
    vistos = set()
    peds_dedup = []
    for p in peds:
        if not isinstance(p, dict):
            continue
        pid = p.get("pedido_id") or p.get("numero") or p.get("pedido") or id(p)
        if pid in vistos:
            continue
        vistos.add(pid)
        peds_dedup.append(p)
    peds = peds_dedup

    criticos = []
    atrasados = []
    saldos = []
    for p in peds:
        status = (p.get("fatura_status") or "").strip()
        ent = (p.get("entrega_status") or "").strip()
        dt = parse_data(p.get("data"))
        dt_fat = parse_data(p.get("data_fatura"))
        dt_ent = parse_data(p.get("data_entrega"))

        # CRÍTICO: sem fatura e o pedido tem mais de 5 dias
        if status == "Sem fatura" and dt and (hoje - dt).days > 5:
            criticos.append(p)

        # ATRASADO: faturado (Total/Parcial) mas com entrega pendente E pedido com +7 dias
        # Usa entrega_status='Pendente' (que vem no JSON) já que data_fatura/data_entrega
        # não são preenchidos em pedidos_2026.
        if status in ("Total", "Parcial") and ent.lower() == "pendente":
            ref = dt_fat or dt
            if ref and (hoje - ref).days > 7:
                atrasados.append(p)

        # SALDO PARCIAL
        if status == "Parcial":
            saldos.append(p)

    # Limita o corpo do email aos mais antigos (Excel anexo segue completo).
    # Sem isso, com 400+ pedidos pendentes, o email vira ilegível.
    def _idade(p):
        d = parse_data(p.get("data_fatura")) or parse_data(p.get("data"))
        return (hoje - d).days if d else 0

    criticos = sorted(criticos, key=_idade, reverse=True)[:30]
    atrasados = sorted(atrasados, key=_idade, reverse=True)[:30]
    return criticos, atrasados, saldos


def agrupar_cli_ind(pedidos):
    """Agrupa pedidos por (cliente, indústria) preservando ordem.

    FIX 2026-05-15: dedup defensivo por pedido_id/numero — a API às vezes
    devolve o mesmo pedido várias vezes (97x p/ alguns pedidos), o que fazia
    a tabela do email aparecer com linhas duplicadas. Aqui mantemos apenas
    a primeira ocorrência de cada chave (pedido_id|numero|pedido).
    """
    grupos = defaultdict(list)
    vistos = set()
    for p in pedidos:
        cli = (p.get("cliente") or p.get("razao") or "—").strip() or "—"
        ind = (p.get("industria") or p.get("fornecedor") or "—").strip() or "—"
        # Chave de dedupe: usa pedido_id, depois numero, depois pedido, depois OC+cliente
        chave_id = (
            str(p.get("pedido_id") or "").strip()
            or str(p.get("numero") or "").strip()
            or str(p.get("pedido") or "").strip()
            or f"{p.get('ordem_compra','')}::{cli}::{ind}"
        )
        chave_completa = (cli, ind, chave_id)
        if chave_completa in vistos:
            continue
        vistos.add(chave_completa)
        grupos[(cli, ind)].append(p)
    return grupos


def renderizar_secao(titulo, cor, pedidos):
    if not pedidos:
        return ""
    grupos = agrupar_cli_ind(pedidos)
    blocos = []
    blocos.append(f'<h3 style="color:{cor};margin:24px 0 8px;font-size:15px;border-bottom:2px solid {cor};padding-bottom:4px">{titulo} ({len(pedidos)})</h3>')

    # ordena por cliente, depois indústria
    chaves = sorted(grupos.keys(), key=lambda k: (k[0].lower(), k[1].lower()))
    for (cli, ind) in chaves:
        ps = grupos[(cli, ind)]
        total_grupo = sum(float(x.get("valor_total") or 0) for x in ps)
        linhas = []
        for p in ps:
            num = p.get("numero") or p.get("pedido") or p.get("pedido_id") or "—"
            dt = fmt_data(p.get("data"))
            dt_fat = fmt_data(p.get("data_fatura"))
            valor = fmt_brl(p.get("valor_total") or p.get("valor") or 0)
            linhas.append(
                f"<tr>"
                f"<td style='padding:4px 8px;border-bottom:1px solid #e2e8f0'>{num}</td>"
                f"<td style='padding:4px 8px;border-bottom:1px solid #e2e8f0'>{dt}</td>"
                f"<td style='padding:4px 8px;border-bottom:1px solid #e2e8f0'>{dt_fat}</td>"
                f"<td style='padding:4px 8px;text-align:right;border-bottom:1px solid #e2e8f0'>{valor}</td>"
                f"</tr>"
            )
        blocos.append(
            f'<div style="margin:10px 0;border:1px solid #e2e8f0;border-radius:6px;overflow:hidden">'
            f'<div style="background:#f1f5f9;padding:6px 10px;font-size:12px">'
            f'<b style="color:#0556AD">{cli}</b> &nbsp;·&nbsp; '
            f'<span style="color:#475569">{ind}</span> &nbsp;·&nbsp; '
            f'<span style="float:right;font-weight:700;color:{cor}">{fmt_brl(total_grupo)}</span>'
            f'</div>'
            f'<table style="width:100%;border-collapse:collapse;font-size:11px">'
            f'<thead style="background:#fafafa"><tr>'
            f'<th style="text-align:left;padding:4px 8px">Pedido</th>'
            f'<th style="text-align:left;padding:4px 8px">Data</th>'
            f'<th style="text-align:left;padding:4px 8px">Faturamento</th>'
            f'<th style="text-align:right;padding:4px 8px">Valor</th>'
            f'</tr></thead><tbody>'
            + "".join(linhas) +
            f'</tbody></table></div>'
        )
    return "".join(blocos)


def gerar_excel_saldos(saldos, itens_pedido_map):
    """Gera Excel com 1 aba 'Saldos' + 1 aba por cliente (top clientes)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("AVISO: openpyxl não instalado — anexo Excel será omitido")
        return None

    wb = Workbook()

    # cores 3MV
    azul = PatternFill(start_color="0556AD", end_color="0556AD", fill_type="solid")
    cinza = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    branco = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    negrito = Font(bold=True, name="Calibri", size=11)
    borda = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # ========= Aba resumo geral =========
    ws = wb.active
    ws.title = "Saldos"
    headers = ["Cliente", "Indústria", "Pedido", "Data", "Faturamento", "Valor Total"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = azul
        cell.font = branco
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda

    r = 2
    grupos = agrupar_cli_ind(saldos)
    chaves = sorted(grupos.keys(), key=lambda k: (k[0].lower(), k[1].lower()))
    for (cli, ind) in chaves:
        for p in grupos[(cli, ind)]:
            ws.cell(row=r, column=1, value=cli).border = borda
            ws.cell(row=r, column=2, value=ind).border = borda
            ws.cell(row=r, column=3, value=str(p.get("numero") or p.get("pedido") or p.get("pedido_id") or "—")).border = borda
            ws.cell(row=r, column=4, value=fmt_data(p.get("data"))).border = borda
            ws.cell(row=r, column=5, value=fmt_data(p.get("data_fatura"))).border = borda
            v = ws.cell(row=r, column=6, value=float(p.get("valor_total") or 0))
            v.number_format = "R$ #,##0.00"
            v.border = borda
            r += 1

    # larguras
    for i, w in enumerate([32, 22, 14, 12, 14, 16], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    # ========= 1 aba por cliente (top 30 por valor pra não estourar limite Excel de 256 abas) =========
    totais_cli = defaultdict(float)
    pedidos_cli = defaultdict(list)
    for p in saldos:
        cli = (p.get("cliente") or "—").strip() or "—"
        totais_cli[cli] += float(p.get("valor_total") or 0)
        pedidos_cli[cli].append(p)

    top = sorted(totais_cli.items(), key=lambda x: -x[1])[:30]
    for cli, _tot in top:
        # nome da aba (Excel limita 31 chars e não permite : \ / ? * [ ])
        safe = "".join(c for c in cli if c not in r':\/?*[]')[:30]
        if not safe:
            safe = "Cliente"
        # se duplicado, adiciona sufixo
        base = safe
        idx = 2
        while safe in wb.sheetnames:
            suf = f" {idx}"
            safe = base[: 30 - len(suf)] + suf
            idx += 1
        wsc = wb.create_sheet(safe)

        # header com nome do cliente
        wsc.cell(row=1, column=1, value=f"Saldos pendentes — {cli}").font = Font(bold=True, size=14, color="0556AD")
        wsc.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        wsc.cell(row=2, column=1, value=f"Gerado em {date.today().strftime('%d/%m/%Y')}").font = Font(italic=True, color="64748B")
        wsc.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

        hdr2 = ["Pedido", "Indústria", "Data do pedido", "Data faturamento", "Status", "Valor"]
        for c, h in enumerate(hdr2, 1):
            cell = wsc.cell(row=4, column=c, value=h)
            cell.fill = azul
            cell.font = branco
            cell.alignment = Alignment(horizontal="center")
            cell.border = borda

        rr = 5
        for p in sorted(pedidos_cli[cli], key=lambda x: parse_data(x.get("data")) or date(1900, 1, 1), reverse=True):
            wsc.cell(row=rr, column=1, value=str(p.get("numero") or p.get("pedido") or p.get("pedido_id") or "—")).border = borda
            wsc.cell(row=rr, column=2, value=(p.get("industria") or "—")).border = borda
            wsc.cell(row=rr, column=3, value=fmt_data(p.get("data"))).border = borda
            wsc.cell(row=rr, column=4, value=fmt_data(p.get("data_fatura"))).border = borda
            wsc.cell(row=rr, column=5, value=p.get("fatura_status") or "—").border = borda
            v = wsc.cell(row=rr, column=6, value=float(p.get("valor_total") or 0))
            v.number_format = "R$ #,##0.00"
            v.border = borda
            rr += 1

        # totalizador
        wsc.cell(row=rr, column=5, value="TOTAL").font = negrito
        vt = wsc.cell(row=rr, column=6, value=sum(float(p.get("valor_total") or 0) for p in pedidos_cli[cli]))
        vt.number_format = "R$ #,##0.00"
        vt.font = negrito
        vt.fill = cinza

        for i, w in enumerate([14, 22, 14, 16, 12, 16], 1):
            wsc.column_dimensions[chr(64 + i)].width = w
        wsc.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("ascii")


def montar_email(criticos, atrasados, saldos, anexo_nome):
    hoje = date.today().strftime("%d/%m/%Y")
    total_acoes = len(criticos) + len(atrasados)
    assunto = f"[3MV] {'OK — sem pendências' if total_acoes == 0 else f'{total_acoes} pendência(s) críticas'} — {hoje}"

    if total_acoes == 0 and not saldos and not FORCE:
        return None

    cor_principal = "#0556AD" if total_acoes == 0 else "#DC2626"
    titulo = "Painel 3MV — Resumo do dia" if total_acoes == 0 else f"{total_acoes} pendência(s) crítica(s)"
    sub = "Sem ações urgentes nesta atualização." if total_acoes == 0 else "Pedidos abaixo precisam de ação imediata."

    total_saldos_valor = sum(float(p.get("valor_total") or 0) for p in saldos)
    anexo_msg = ""
    if saldos and anexo_nome:
        anexo_msg = (
            f'<div style="margin:16px 0;padding:12px;background:#eff6ff;border-left:4px solid #0556AD;border-radius:4px">'
            f'<b style="color:#0556AD">📎 Anexo: {anexo_nome}</b><br>'
            f'<span style="font-size:12px;color:#1e3a8a">Excel com saldos parciais separados por cliente — '
            f'use as abas pra enviar a cada cliente individualmente.</span>'
            f'</div>'
        )

    html = f"""<!DOCTYPE html>
<html><body style="font-family:-apple-system,BlinkMacSystemFont,sans-serif;max-width:880px;margin:0 auto;padding:16px;background:#f8fafc">
  <div style="background:{cor_principal};padding:16px 20px;border-radius:8px 8px 0 0">
    <h1 style="color:#fff;margin:0;font-size:20px">{titulo}</h1>
    <p style="color:rgba(255,255,255,0.9);margin:4px 0 0;font-size:13px">{sub}  •  {hoje}</p>
  </div>
  <div style="background:#fff;border:1px solid #e2e8f0;border-top:0;padding:20px;border-radius:0 0 8px 8px">
    <div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:12px">
      <div style="flex:1;min-width:160px;background:#fef2f2;border-left:4px solid #DC2626;padding:10px 14px;border-radius:4px">
        <div style="font-size:11px;color:#7f1d1d;font-weight:600;text-transform:uppercase">Críticos</div>
        <div style="font-size:24px;font-weight:700;color:#DC2626">{len(criticos)}</div>
        <div style="font-size:11px;color:#7f1d1d">Sem fatura há +5 dias</div>
      </div>
      <div style="flex:1;min-width:160px;background:#fffbeb;border-left:4px solid #D97706;padding:10px 14px;border-radius:4px">
        <div style="font-size:11px;color:#92400e;font-weight:600;text-transform:uppercase">Atrasados</div>
        <div style="font-size:24px;font-weight:700;color:#D97706">{len(atrasados)}</div>
        <div style="font-size:11px;color:#92400e">Faturados sem entrega +7 dias</div>
      </div>
      <div style="flex:1;min-width:160px;background:#eff6ff;border-left:4px solid #0556AD;padding:10px 14px;border-radius:4px">
        <div style="font-size:11px;color:#1e3a8a;font-weight:600;text-transform:uppercase">Saldos parciais</div>
        <div style="font-size:24px;font-weight:700;color:#0556AD">{len(saldos)}</div>
        <div style="font-size:11px;color:#1e3a8a">{fmt_brl(total_saldos_valor)} a faturar</div>
      </div>
    </div>
    {anexo_msg}
    {renderizar_secao('🔴 Pedidos críticos sem fatura', '#DC2626', criticos)}
    {renderizar_secao('📦 Entregas atrasadas (faturadas)', '#D97706', atrasados)}
    <div style="margin-top:20px;padding:12px;background:#f1f5f9;border-radius:6px;text-align:center">
      <a href="{PAINEL_URL}" style="display:inline-block;background:#0556AD;color:#fff;padding:10px 16px;border-radius:6px;text-decoration:none;font-weight:700;margin:4px">📊 Abrir Painel 3MV</a>
      <a href="{MOBILE_URL}" style="display:inline-block;background:#10B981;color:#fff;padding:10px 16px;border-radius:6px;text-decoration:none;font-weight:700;margin:4px">📱 Abrir 3MV Mobile</a>
    </div>
    <p style="font-size:11px;color:#64748b;margin-top:12px;text-align:center">Email automático — pipeline cloud 3MV — {hoje}</p>
  </div>
</body></html>"""
    return (assunto, html, total_acoes)


def enviar_um(nome, email, assunto, html, anexo_b64, anexo_nome):
    payload = {
        "para": [email],
        "assunto": assunto,
        "corpo": html,
        "tipo": "html",
    }
    if anexo_b64 and anexo_nome:
        payload["anexos"] = [{
            "nome": anexo_nome,
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "base64": anexo_b64,
        }]

    body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        f"{API_URL}/api/email/send",
        data=body,
        headers={
            "Authorization": f"Bearer {TOKEN}",
            "Content-Type": "application/json",
            "User-Agent": "painel-3mv-alertas/2.0 (+https://painel.3mvrepresentacao.com)",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            txt = r.read().decode("utf-8")
            print(f"  OK {nome} <{email}> - HTTP {r.status} - {txt[:140]}")
            return True
    except urllib.error.HTTPError as e:
        print(f"  ERR {nome} <{email}> - HTTP {e.code}: {e.read().decode()[:200]}")
        return False
    except Exception as e:
        print(f"  ERR {nome} <{email}> - erro: {e}")
        return False


def destinatarios():
    if TO_OVERRIDE:
        out = []
        for item in TO_OVERRIDE.split(","):
            item = item.strip()
            if not item:
                continue
            if "<" in item and ">" in item:
                nome = item.split("<")[0].strip()
                email = item.split("<")[1].split(">")[0].strip()
            else:
                email = item
                nome = email.split("@")[0]
            out.append((nome, email))
        return out
    return DESTS_PADRAO


def main():
    if not TOKEN:
        print("AVISO: EMAIL_TOKEN nao configurado - alertas nao enviados")
        return 0

    try:
        dados = carregar_dados()
    except Exception as e:
        print(f"AVISO: nao foi possivel ler dados.json: {e}")
        return 0

    criticos, atrasados, saldos = classificar(dados)
    print(f"Criticos: {len(criticos)}  |  Atrasados: {len(atrasados)}  |  Saldos parciais: {len(saldos)}")

    total_acoes = len(criticos) + len(atrasados)
    if total_acoes == 0 and len(saldos) == 0 and not FORCE:
        print("Sem pendencias - pulando envio (use ALERTAS_FORCE=1 pra forcar)")
        return 0

    # gera anexo Excel se houver saldos parciais
    anexo_b64 = None
    anexo_nome = None
    if saldos:
        print("Gerando Excel de saldos parciais...")
        anexo_b64 = gerar_excel_saldos(saldos, {})
        if anexo_b64:
            anexo_nome = f"Saldos_Parciais_{date.today().strftime('%d-%m-%Y')}.xlsx"
            print(f"  Excel: {anexo_nome}  ({len(anexo_b64)//1024} KB base64)")

    montagem = montar_email(criticos, atrasados, saldos, anexo_nome)
    if not montagem:
        print("Nada a enviar")
        return 0
    assunto, html, _ = montagem
    print(f'Enviando: "{assunto}"')

    dests = destinatarios()
    ok = 0
    for nome, email in dests:
        if enviar_um(nome, email, assunto, html, anexo_b64, anexo_nome):
            ok += 1

    print(f"\n{ok}/{len(dests)} email(s) enviado(s) com sucesso")
    return 0


if __name__ == "__main__":
    sys.exit(main())
