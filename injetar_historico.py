#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Injeta histórico 2024/2025/2026 + agregações mensais detalhadas no dados.json
para popular Dashboard, Comparativo 25×26, Top Clientes, Top Indústrias, Top Mensal.
"""
import json
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict
try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install',
                            '--break-system-packages', '--quiet', 'openpyxl'])
    from openpyxl import load_workbook


def _gd_base():
    candidates = [
        Path.home() / 'Library/CloudStorage/GoogleDrive-marcelo@3mvrepresentacao.com/Meu Drive',
    ]
    for c in candidates:
        if (c / '3MV/BI/2026/pedidos 2024_2025_2026.xlsx').exists():
            return c
    return candidates[0]


GD = _gd_base()
HIST_XLSX = GD / '3MV/BI/2026/pedidos 2024_2025_2026.xlsx'
DADOS_JSON = Path(__file__).parent / 'dados.json'

# Mapeamento: indústrias renomeadas/incorporadas — fundem o histórico com a marca atual.
# Cada regra: lista de palavras-chave → nome canônico final.
INDUSTRIAS_FUSAO = [
    # M Dias Branco absorveu 55F Moinho Isabela; também unifica truncamentos
    (['m dias branco', '55f', 'moinho isabela'], 'M DIAS BRANCO'),
    # Ruby Rose nasceu da Union Medic
    (['ruby rose', 'union medic'], 'RUBY ROSE'),
]

def normaliza_industria(nome):
    """Normaliza nome de indústria — funde marcas renomeadas/truncadas em um nome canônico."""
    if not nome:
        return '—'
    n = str(nome).strip()
    n_lower = n.lower()
    for keywords, canonico in INDUSTRIAS_FUSAO:
        if any(k in n_lower for k in keywords):
            return canonico
    return n


def main():
    if not HIST_XLSX.exists():
        print(f"⚠ {HIST_XLSX} não existe — pulando", file=sys.stderr)
        return 0
    if not DADOS_JSON.exists():
        print(f"✗ {DADOS_JSON} não existe", file=sys.stderr)
        return 1

    print(f"→ Lendo {HIST_XLSX}")
    wb = load_workbook(str(HIST_XLSX), data_only=True)
    ws = wb['Pedidos']

    por_ind        = defaultdict(lambda: defaultdict(float))   # {ind: {ano: vlr}}
    por_cli        = defaultdict(lambda: defaultdict(float))   # {cli: {ano: vlr}}
    por_ano        = defaultdict(float)
    # Agregações mensais ano-mes
    por_mes_ano    = defaultdict(lambda: defaultdict(float))   # {ano: {mes: vlr}}
    por_cli_mes_25 = defaultdict(lambda: defaultdict(float))   # {cli: {mes 1-12: vlr}} 2025
    por_cli_mes_26 = defaultdict(lambda: defaultdict(float))
    por_ind_mes_25 = defaultdict(lambda: defaultdict(float))
    por_ind_mes_26 = defaultdict(lambda: defaultdict(float))
    count = 0

    # 2024 e 2025: lê do XLSX (histórico consolidado do Power BI)
    # 2026: vem da API (pedidos_2026), abaixo — sempre atualizado
    for r in range(2, ws.max_row + 1):
        cliente   = ws.cell(row=r, column=2).value
        industria = ws.cell(row=r, column=3).value
        valor_fat = ws.cell(row=r, column=8).value or 0
        dt_venda  = ws.cell(row=r, column=10).value
        if not isinstance(valor_fat, (int, float)) or valor_fat <= 0:
            continue
        ano, mes = None, None
        if isinstance(dt_venda, datetime):
            ano, mes = dt_venda.year, dt_venda.month
        elif isinstance(dt_venda, str) and len(dt_venda) >= 7:
            try:
                ano, mes = int(dt_venda[:4]), int(dt_venda[5:7])
            except: continue
        if not ano or ano not in (2024, 2025):
            continue  # 2026 vem da API, ignora XLSX

        ind_n = normaliza_industria(industria)
        cli_n = str(cliente).strip() if cliente else '—'

        por_ind[ind_n][ano] += valor_fat
        por_cli[cli_n][ano] += valor_fat
        por_ano[ano] += valor_fat
        por_mes_ano[ano][mes] += valor_fat

        if ano == 2025:
            por_cli_mes_25[cli_n][mes] += valor_fat
            por_ind_mes_25[ind_n][mes] += valor_fat
        count += 1

    # 2026: usa pedidos_2026 da API (sempre atualizado, mar/abr inclusos)
    print(f"→ Lendo pedidos_2026 da API (dados.json)")
    d_api = json.loads(DADOS_JSON.read_text(encoding='utf-8'))
    peds_2026 = d_api.get('pedidos_2026', [])
    count_2026 = 0
    for p in peds_2026:
        valor = p.get('valor_total', 0) or 0
        if not isinstance(valor, (int, float)) or valor <= 0:
            continue
        dt = p.get('data', '') or ''
        mes = None
        if dt and len(dt) >= 7:
            try:
                ano_p = int(dt[:4]); mes = int(dt[5:7])
                if ano_p != 2026: continue
            except: continue
        if not mes:
            continue
        cli_n = str(p.get('cliente') or '—').strip()
        ind_n = normaliza_industria(p.get('industria'))
        por_ind[ind_n][2026] += valor
        por_cli[cli_n][2026] += valor
        por_ano[2026] += valor
        por_mes_ano[2026][mes] += valor
        por_cli_mes_26[cli_n][mes] += valor
        por_ind_mes_26[ind_n][mes] += valor
        count_2026 += 1
    print(f"  2026 da API: {count_2026} pedidos · R$ {por_ano[2026]:,.2f}")
    count += count_2026

    def calc_cresc(d):
        out = {}
        for nome, anos in d.items():
            v24 = anos.get(2024, 0); v25 = anos.get(2025, 0); v26 = anos.get(2026, 0)
            cresc_25 = (v25 / v24 - 1) if v24 > 0 else None
            cresc_26 = (v26 / v25 - 1) if v25 > 0 else None
            out[nome] = {
                '2024': round(v24, 2), '2025': round(v25, 2), '2026': round(v26, 2),
                'cresc_25': round(cresc_25, 4) if cresc_25 is not None else None,
                'cresc_26': round(cresc_26, 4) if cresc_26 is not None else None,
            }
        return out

    def mes_dict_to_list(d):
        """Converte {nome: {1..12: vlr}} → {nome: [12 valores]}"""
        out = {}
        for nome, meses in d.items():
            out[nome] = [round(meses.get(m, 0), 2) for m in range(1, 13)]
        return out

    historico = {
        'por_ano': {str(k): round(v, 2) for k, v in por_ano.items()},
        'por_industria': calc_cresc(por_ind),
        'por_cliente':   calc_cresc(por_cli),
        'mensal_25': [round(por_mes_ano[2025].get(m, 0), 2) for m in range(1, 13)],
        'mensal_26': [round(por_mes_ano[2026].get(m, 0), 2) for m in range(1, 13)],
        'cli_mes_25': mes_dict_to_list(por_cli_mes_25),
        'cli_mes_26': mes_dict_to_list(por_cli_mes_26),
        'ind_mes_25': mes_dict_to_list(por_ind_mes_25),
        'ind_mes_26': mes_dict_to_list(por_ind_mes_26),
        'count_pedidos': count,
        'data_referencia': datetime.now().strftime('%Y-%m-%d %H:%M'),
    }

    print(f"  Pedidos: {count} | "
          f"24=R${por_ano[2024]:,.0f} 25=R${por_ano[2025]:,.0f} 26=R${por_ano[2026]:,.0f}")
    print(f"  Indústrias: {len(por_ind)} | Clientes: {len(por_cli)}")

    d = json.loads(DADOS_JSON.read_text(encoding='utf-8'))
    d['historico'] = historico
    DADOS_JSON.write_text(json.dumps(d, ensure_ascii=False), encoding='utf-8')
    print(f"✓ dados.json atualizado")
    return 0


if __name__ == '__main__':
    sys.exit(main())
