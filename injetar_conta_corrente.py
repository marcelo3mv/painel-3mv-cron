#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Injeta dados de /3MV/BI/2026/ContaCorrente.xlsx no dados.json do painel.

3 abas da planilha:
  • NF DE DEV EM ABERTO  → seção 'nfd_aberto'  (a receber)
  • NF DE DEV PAGA       → seção 'nfd_paga'    (já recebido)
  • CONTA CORRENTE       → seção 'conta_corrente' (negociações)

Adiciona ao dados.json:
  d['conta_corrente'] = {
    'nfd_aberto':  { 'total': float, 'count': int, 'itens': [...] },
    'nfd_paga':    { 'total': float, 'count': int, 'itens': [...] },
    'cc_pedidos':  { 'total': float, 'pago': float, 'saldo': float, 'count': int, 'itens': [...] },
    'data_referencia': '2026-05-06',
  }
"""
import json
import sys
from pathlib import Path
from datetime import datetime
try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install',
                            '--break-system-packages', '--quiet', 'openpyxl'])
    from openpyxl import load_workbook


def _gd_base():
    candidates = [
        Path.home() / 'Library/CloudStorage/GoogleDrive-marcelo@3mvrepresentacao.com/Meu Drive',
    ]
    for c in candidates:
        if (c / '3MV/BI/2026/ContaCorrente.xlsx').exists():
            return c
    return candidates[0]


GD = _gd_base()
CC_XLSX = GD / '3MV/BI/2026/ContaCorrente.xlsx'
DADOS_JSON = Path(__file__).parent / 'dados.json'


def _val(cell):
    """Retorna valor de célula limpo."""
    v = cell.value
    if v is None: return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)):
        return v
    return str(v).strip()


def ler_nfd_aberto(wb):
    """NF DE DEV EM ABERTO: A=Industria | B=N NFD | C=Data | D=Cliente |
       E=Indústria origem | F=CNPJ | G=Valor | H=Obs | I=NF Origem"""
    if 'NF DE DEV EM ABERTO' not in wb.sheetnames:
        return {'total': 0, 'count': 0, 'itens': []}
    ws = wb['NF DE DEV EM ABERTO']
    itens = []
    total = 0.0
    for r in range(2, ws.max_row + 1):
        valor = _val(ws.cell(row=r, column=7))   # G
        if not isinstance(valor, (int, float)) or valor <= 0:
            continue
        item = {
            'industria_emissora': _val(ws.cell(row=r, column=1)),
            'nfd':                _val(ws.cell(row=r, column=2)),
            'data':               _val(ws.cell(row=r, column=3)),
            'cliente':            _val(ws.cell(row=r, column=4)),
            'industria_origem':   _val(ws.cell(row=r, column=5)),
            'cnpj':               _val(ws.cell(row=r, column=6)),
            'valor':              float(valor),
            'observacao':         _val(ws.cell(row=r, column=8)),
            'nf_origem':          _val(ws.cell(row=r, column=9)),
        }
        itens.append(item)
        total += float(valor)
    return {'total': total, 'count': len(itens), 'itens': itens}


def ler_nfd_paga(wb):
    """NF DE DEV PAGA: A=Nº NFD | B=Cliente | C=Indústria | D=Pago na NF |
       E=Data NFD | F=Valor | G=Obs | H=NF Origem"""
    if 'NF DE DEV PAGA' not in wb.sheetnames:
        return {'total': 0, 'count': 0, 'itens': []}
    ws = wb['NF DE DEV PAGA']
    itens = []
    total = 0.0
    for r in range(2, ws.max_row + 1):
        # Tenta col F (valor); senão col D (Pago na NF) também é valor
        v_f = _val(ws.cell(row=r, column=6))
        v_d = _val(ws.cell(row=r, column=4))
        valor = v_f if isinstance(v_f, (int, float)) and v_f > 0 else (
                v_d if isinstance(v_d, (int, float)) and v_d > 0 else 0)
        if not valor or valor <= 0:
            continue
        item = {
            'nfd':              _val(ws.cell(row=r, column=1)),
            'cliente':          _val(ws.cell(row=r, column=2)),
            'industria':        _val(ws.cell(row=r, column=3)),
            'pago_na_nf':       v_d if isinstance(v_d, (int, float)) else 0,
            'data':             _val(ws.cell(row=r, column=5)),
            'valor':            float(valor),
            'observacao':       _val(ws.cell(row=r, column=7)),
            'nf_origem':        _val(ws.cell(row=r, column=8)),
        }
        itens.append(item)
        total += float(valor)
    return {'total': total, 'count': len(itens), 'itens': itens}


def ler_conta_corrente(wb):
    """CONTA CORRENTE: A=Pedido | B=Cliente | D=Indústria | E=Tipo |
       F=Vlr Total Pedido E | G=Vlr Total Pedido F | H=% | I=Vlr Negociado |
       J=Vlr Negociado pelo | K=Vlr Pago | L=Saldo | M=NF | N=Data"""
    if 'CONTA CORRENTE' not in wb.sheetnames:
        return {'total': 0, 'pago': 0, 'saldo': 0, 'count': 0, 'itens': []}
    ws = wb['CONTA CORRENTE']
    itens = []
    total_negociado = 0.0
    total_pago = 0.0
    total_saldo = 0.0
    for r in range(2, ws.max_row + 1):
        cliente = _val(ws.cell(row=r, column=2))
        if not cliente:
            continue
        vlr_negociado = _val(ws.cell(row=r, column=9))   # I
        vlr_pago      = _val(ws.cell(row=r, column=11))  # K
        saldo         = _val(ws.cell(row=r, column=12))  # L
        if not isinstance(vlr_negociado, (int, float)):
            vlr_negociado = 0
        if not isinstance(vlr_pago, (int, float)):
            vlr_pago = 0
        if not isinstance(saldo, (int, float)):
            saldo = 0
        if vlr_negociado <= 0 and saldo <= 0 and vlr_pago <= 0:
            continue
        item = {
            'pedido':           _val(ws.cell(row=r, column=1)),
            'cliente':          cliente,
            'negociacao':       _val(ws.cell(row=r, column=3)),
            'industria':        _val(ws.cell(row=r, column=4)),
            'tipo':             _val(ws.cell(row=r, column=5)),
            'vlr_total':        _val(ws.cell(row=r, column=6)) or 0,
            'pct':              _val(ws.cell(row=r, column=8)) or 0,
            'vlr_negociado':    float(vlr_negociado),
            'vlr_pago':         float(vlr_pago),
            'saldo':            float(saldo),
            'nf':               _val(ws.cell(row=r, column=13)),
            'data':             _val(ws.cell(row=r, column=14)),
        }
        itens.append(item)
        total_negociado += float(vlr_negociado)
        total_pago += float(vlr_pago)
        total_saldo += float(saldo)
    return {
        'total': total_negociado,
        'pago': total_pago,
        'saldo': total_saldo,
        'count': len(itens),
        'itens': itens,
    }



def carregar_prazos():
    """Carrega prazos_clientes.json se existir."""
    p = Path(__file__).parent / 'prazos_clientes.json'
    if not p.exists():
        return {'_default_geral': 30, 'prazos': {}}
    return json.loads(p.read_text(encoding='utf-8'))


def prazo_do_cliente(nome_cliente, prazos_cfg):
    """Retorna o prazo (dias) configurado pra um cliente, ou _default_geral."""
    if not nome_cliente:
        return prazos_cfg.get('_default_geral', 30)
    nome = str(nome_cliente).strip().upper()
    prazos = prazos_cfg.get('prazos', {})
    # Busca exata
    if nome in prazos:
        return prazos[nome]
    # Busca parcial (caso seja "INDIANA CD BH" e eu tenha só "INDIANA")
    for cli_nome, dias in prazos.items():
        if cli_nome.upper() in nome:
            return dias
    return prazos_cfg.get('_default_geral', 30)


def enriquecer_nfd_com_prazos(itens_aberto, prazos_cfg):
    """Para cada NFD em aberto, adiciona prazo_dias, dias_decorridos, dias_para_vencer, status_atraso."""
    from datetime import datetime, date
    hoje = date.today()
    for item in itens_aberto:
        cliente = item.get('cliente') or ''
        prazo = prazo_do_cliente(cliente, prazos_cfg)
        item['prazo_dias'] = prazo

        data_str = item.get('data')
        if data_str and isinstance(data_str, str) and len(data_str) >= 10:
            try:
                d = datetime.strptime(data_str[:10], '%Y-%m-%d').date()
                decorridos = (hoje - d).days
                vencer_em = prazo - decorridos
                item['dias_decorridos'] = decorridos
                item['dias_para_vencer'] = vencer_em
                if decorridos > prazo:
                    dias_atraso = decorridos - prazo
                    item['status_atraso'] = 'atrasado'
                    item['dias_atraso'] = dias_atraso
                elif decorridos > prazo * 0.8:  # > 80% do prazo
                    item['status_atraso'] = 'proximo_vencimento'
                else:
                    item['status_atraso'] = 'em_dia'
            except Exception:
                item['status_atraso'] = 'sem_data'
        else:
            item['status_atraso'] = 'sem_data'
    return itens_aberto


def main():
    if not CC_XLSX.exists():
        print(f"⚠ {CC_XLSX} não existe — pulando injeção", file=sys.stderr)
        return 0
    if not DADOS_JSON.exists():
        print(f"✗ {DADOS_JSON} não existe — extrair_saldos precisa rodar antes", file=sys.stderr)
        return 1

    print(f"→ Lendo {CC_XLSX}")
    wb = load_workbook(str(CC_XLSX), data_only=True)

    nfd_aberto = ler_nfd_aberto(wb)
    prazos_cfg = carregar_prazos()
    nfd_aberto['itens'] = enriquecer_nfd_com_prazos(nfd_aberto.get('itens', []), prazos_cfg)
    # Calcula totais por status
    totais_atraso = {'atrasado': 0, 'proximo_vencimento': 0, 'em_dia': 0, 'sem_data': 0}
    for it in nfd_aberto['itens']:
        st = it.get('status_atraso', 'sem_data')
        totais_atraso[st] = totais_atraso.get(st, 0) + (it.get('valor', 0) or 0)
    nfd_aberto['por_status'] = totais_atraso
    nfd_paga   = ler_nfd_paga(wb)
    cc         = ler_conta_corrente(wb)

    print(f"  NF DEV EM ABERTO: {nfd_aberto['count']} NFDs · R$ {nfd_aberto['total']:,.2f}")
    print(f"  NF DEV PAGA:      {nfd_paga['count']} NFDs · R$ {nfd_paga['total']:,.2f}")
    print(f"  CONTA CORRENTE:   {cc['count']} pedidos · "
          f"Negociado R$ {cc['total']:,.2f} · "
          f"Pago R$ {cc['pago']:,.2f} · Saldo R$ {cc['saldo']:,.2f}")

    print(f"→ Lendo {DADOS_JSON}")
    d = json.loads(DADOS_JSON.read_text(encoding='utf-8'))

    # FASE 9.4: Captura mtime da planilha BI pra mostrar no painel
    try:
        planilha_mtime = datetime.fromtimestamp(CC_XLSX.stat().st_mtime).strftime('%Y-%m-%d %H:%M')
    except Exception:
        planilha_mtime = ''

    d['conta_corrente'] = {
        'nfd_aberto':     nfd_aberto,
        'nfd_paga':       nfd_paga,
        'cc_pedidos':     cc,
        'data_referencia': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'planilha_mtime': planilha_mtime,
        'planilha_path': str(CC_XLSX),
    }

    DADOS_JSON.write_text(
        json.dumps(d, ensure_ascii=False),
        encoding='utf-8',
    )
    print(f"✓ dados.json atualizado com seção 'conta_corrente'")
    return 0


if __name__ == '__main__':
    sys.exit(main())
