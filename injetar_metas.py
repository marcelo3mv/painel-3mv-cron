#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Injeta dados de /3MV/BI/2026/Metas.xlsx no dados.json do painel.

2 abas da planilha:
  • META GERAL      → meta_geral_mensal {1..12: valor}
  • META INDUSTRIA  → meta_por_industria_mensal {industria: {1..12: valor}}

Adiciona ao dados.json:
  d['metas'] = {
    'ano': 2026,
    'meta_geral_mensal': {...},
    'meta_por_industria_mensal': {...},
    'planilha_mtime': 'YYYY-MM-DD HH:MM',
    'planilha_path': '/path/to/Metas.xlsx',
  }

API Suas Vendas NÃO tem endpoint de metas (verificado: /Meta, /MetaVenda,
/MetaIndustria, /MetaMensal, /Objetivo, /Forecast → todos 404).
Por isso usamos planilha BI editável manualmente.
"""
import json
import sys
from pathlib import Path
from datetime import datetime
try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install',
                            '--break-system-packages', '--quiet', 'openpyxl'])
    from openpyxl import load_workbook


def _gd_base():
    candidates = [
        Path.home() / 'Library/CloudStorage/GoogleDrive-marcelo@3mvrepresentacao.com/Meu Drive',
    ]
    for c in candidates:
        if (c / '3MV/BI/2026/Metas.xlsx').exists():
            return c
    return candidates[0]


GD = _gd_base()

# Fallback para GitHub Actions: usar planilhas-snapshot/ do repo se Drive nao existe
_SNAPSHOT = Path(__file__).parent / 'planilhas-snapshot' / 'Metas.xlsx'
METAS_XLSX = GD / '3MV/BI/2026/Metas.xlsx'
if not METAS_XLSX.exists() and _SNAPSHOT.exists():
    METAS_XLSX = _SNAPSHOT
    print(f'  -> usando snapshot do repo: {METAS_XLSX}')

DADOS_JSON = Path(__file__).parent / 'dados.json'

# Mapa MES (string) → número
MES_MAP = {
    'JAN':1, 'FEV':2, 'MAR':3, 'ABR':4, 'MAI':5, 'JUN':6,
    'JUL':7, 'AGO':8, 'SET':9, 'OUT':10, 'NOV':11, 'DEZ':12,
    'JANEIRO':1,'FEVEREIRO':2,'MARÇO':3,'MARCO':3,'ABRIL':4,'MAIO':5,'JUNHO':6,
    'JULHO':7,'AGOSTO':8,'SETEMBRO':9,'OUTUBRO':10,'NOVEMBRO':11,'DEZEMBRO':12,
}


def _to_num(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return float(v)
    try:
        return float(str(v).replace(',', '.').strip())
    except Exception:
        return 0


def _norm_mes(v):
    if v is None: return None
    if isinstance(v, (int, float)):
        n = int(v)
        return n if 1 <= n <= 12 else None
    s = str(v).upper().strip()
    return MES_MAP.get(s)


def ler_meta_geral(wb):
    """META GERAL: A=MÊS, B=ANO, C=META R$"""
    sheet = None
    for sn in wb.sheetnames:
        if 'GERAL' in sn.upper():
            sheet = wb[sn]; break
    if not sheet:
        print('  ⚠ Aba META GERAL não encontrada')
        return {}
    metas = {}
    for r, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]: continue
        mes = _norm_mes(row[0])
        if mes is None: continue
        valor = _to_num(row[2]) if len(row) >= 3 else 0
        metas[str(mes)] = valor
    return metas


def ler_meta_industria(wb):
    """META INDUSTRIA: A=INDUSTRIA, B=MES, C=ANO, D=META R$"""
    sheet = None
    for sn in wb.sheetnames:
        if 'INDUSTRIA' in sn.upper().replace('Ú','U') or 'IND' in sn.upper():
            if 'INSTR' not in sn.upper() and 'INSTRUC' not in sn.upper():
                sheet = wb[sn]; break
    if not sheet:
        print('  ⚠ Aba META INDUSTRIA não encontrada')
        return {}
    metas = {}  # {industria: {mes: valor}}
    for r, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]: continue
        ind = str(row[0]).strip()
        mes = _norm_mes(row[1]) if len(row) >= 2 else None
        valor = _to_num(row[3]) if len(row) >= 4 else 0
        if not ind or mes is None: continue
        if ind not in metas:
            metas[ind] = {}
        metas[ind][str(mes)] = valor
    # Compatibilidade com painel atual: também salva como dict simples (média mensal por indústria)
    metas_simples = {}
    for ind, by_mes in metas.items():
        # Usa média ou primeiro mês como representativo
        vals = list(by_mes.values())
        metas_simples[ind] = sum(vals) / len(vals) if vals else 0
    return metas, metas_simples


def main():
    print(f"\n→ Lendo {METAS_XLSX}")
    if not METAS_XLSX.exists():
        print(f'  ⚠ {METAS_XLSX} não existe — criar via planilha modelo. Pulando injeção.')
        return 0

    wb = load_workbook(METAS_XLSX, data_only=True)
    print(f'  Abas: {wb.sheetnames}')

    meta_geral = ler_meta_geral(wb)
    meta_ind_mensal, meta_ind_simples = ler_meta_industria(wb)
    print(f"  Meta geral: {len(meta_geral)} meses · meta total ano: R$ {sum(meta_geral.values()):,.0f}")
    print(f"  Meta indústria: {len(meta_ind_mensal)} indústrias")

    if not DADOS_JSON.exists():
        print(f"  ⚠ dados.json não existe ({DADOS_JSON}) — pulando injeção")
        return 0
    d = json.loads(DADOS_JSON.read_text(encoding='utf-8'))

    try:
        planilha_mtime = datetime.fromtimestamp(METAS_XLSX.stat().st_mtime).strftime('%Y-%m-%d %H:%M')
    except Exception:
        planilha_mtime = ''

    d['metas'] = {
        '_doc': 'Lido de /3MV/BI/2026/Metas.xlsx — edite a planilha pra atualizar',
        '_atualizado_em': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'ano': 2026,
        'meta_geral_mensal': meta_geral,
        'meta_por_industria_mensal': meta_ind_simples,  # média mensal — compat com renderMetasDashboard
        'meta_por_industria_mensal_detalhado': meta_ind_mensal,  # mês a mês detalhado
        'planilha_mtime': planilha_mtime,
        'planilha_path': str(METAS_XLSX),
    }

    DADOS_JSON.write_text(
        json.dumps(d, ensure_ascii=False),
        encoding='utf-8',
    )
    print(f"✓ dados.json atualizado com seção 'metas'")
    return 0


if __name__ == '__main__':
    sys.exit(main() or 0)
